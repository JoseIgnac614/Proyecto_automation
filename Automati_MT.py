from selenium import webdriver
from selenium.webdriver.common.by import By  # Importa el módulo By
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from datetime import datetime
import openpyxl
import time
import sys
import io
import re
import ctypes
import os
import pygame
import traceback

pestañapredio  =        True                #DECIDE SI LLENAR PESTAÑA PREDIO O NO    

llenarsolopredio =      1                  #LLENAR SOLO INFO DEL PREDIO O TAMBIEN EL RESTO DEL PROCESO????????     0 para solo predio, 1 para todo
meterleservi =          True                #METERLE SERVIDUMBRE O NO???????????????    False para no, True para sí
revisar_interesados =   True                #REVISAR INTERESADOSSS????????????? (Mirar cuales estan mal, eliminar, modificar)
llenar_derechos     =   True
modoqc              =   False                #Modidica interesados pero debe estar revisar_interesados activa, derecho se modifica,
poneraprobado       =   False                #Poner estado en aprobado, False = lo pone en realizado
HibernarPC =            False               #HIBERNAR PC AL TEMRINAR????????

# Abre el archivo Excel
#carpeta_almacenamiento= 'C:/Users/nacho/Downloads/davud/Autofinal/09-11-2023/'
carpeta_almacenamiento = "C:/Users/nacho/Downloads/Pruebas_autom/15-12-2023/"
nombre_excel = 'Libro1.xlsx'
indice_folio = "303-"





nivel_de_zoom = 0.8
# Configura las opciones del navegador para ajustar el nivel de zoom
options = webdriver.ChromeOptions()
options.add_argument(f"--force-device-scale-factor={nivel_de_zoom}")

# Crea una instancia de un navegador web (por ejemplo, Edge)
driver = webdriver.Chrome(options=options)
# Maximiza la ventana del navegador
driver.maximize_window()
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# Evitar que el PC se suspenda
ctypes.windll.kernel32.SetThreadExecutionState(0x80000002)



archivo_excel = openpyxl.load_workbook(carpeta_almacenamiento+nombre_excel)

# Selecciona la hoja en la que deseas trabajar
hoja = archivo_excel['informacion_propiedades']
hojajuridico = archivo_excel['nombres_cedulas']

def reproducir_audio(ruta_audio):
    pygame.init()
    pygame.mixer.init()
    pygame.mixer.music.load(ruta_audio)
    pygame.mixer.music.play()

ruta_audio = "Alarmas/Audio1.mp3"  # Reemplaza con la ruta de tu archivo .mp3

def search_element_click(edito):
    driver.execute_script("arguments[0].scrollIntoView();", edito)   
    time.sleep(.5)     
    
    encontrado = False
    
    while not encontrado:
        try:
            # Intenta encontrar el elemento
            edito.click()
            encontrado = True
        except:
            # Si no se encuentra el elemento, desplaza el scroll hacia arriba
            driver.execute_script("window.scrollBy(0, -100);")
            # Espera un momento para que la página se cargue y se actualice
            time.sleep(.7)  
    time.sleep(.5)    
    
    
def quitar_acentos(texto):
    acentos = {
        'á': 'a', 'é': 'e', 'í': 'i', 'ó': 'o', 'ú': 'u',
        'Á': 'A', 'É': 'E', 'Í': 'I', 'Ó': 'O', 'Ú': 'U'
    }
    
    # Reemplaza caracteres acentuados
    for acento, sin_acento in acentos.items():
        texto = texto.replace(acento, sin_acento)
    
    return texto

def crear_servidumbre(cadenaserv,cadena_escritura,cadena_area):
    del_every_serv = True
    
    
    
    
    wait = WebDriverWait(driver, 15)
    if cadenaserv != "NO":
        time.sleep(1)
        wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "#Tab_TAB1Containerpanel1"))).click()
        time.sleep(1)
        
        checkbox = driver.find_element(By.CSS_SELECTOR, "#W0014vSERVIDUMBRE")
        
        if checkbox.is_selected():
            # checkbox = driver.find_element(By.CSS_SELECTOR, "#W0014vSERVIDUMBRE")      # BOTON DE QUE NO HAY AREA CATASTRAL
            checkbox.click()
            time.sleep(1)
            driver.find_element(By.CSS_SELECTOR, "#W0014GUARDAR").click()
            time.sleep(2)
            wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, "body > div.gx-mask")))
            wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, "body > div.gx-mask.gx-unmask")))
            # time.sleep(2)

        checkbox = driver.find_element(By.CSS_SELECTOR, "#W0014vSERVIDUMBRE")      # BOTON DE QUE NO HAY AREA CATASTRAL
        if not checkbox.is_selected():            
            checkbox.click()
            time.sleep(1)
            driver.find_element(By.CSS_SELECTOR, "#W0014GUARDAR").click()
            time.sleep(2)
            wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, "body > div.gx-mask")))
            wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, "body > div.gx-mask.gx-unmask")))
        
        wait = WebDriverWait(driver, 15)
        panel = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "#Tab_TAB1Containerpanel5")))
        panel.click()
        time.sleep(1.5)
        wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "#W0046vACTSERVTIPOID")))                        
        
        
        
        select_element = driver.find_element(By.CSS_SELECTOR,"#W0046vACTSERVTIPOID")  # Reemplaza el ID con el de tu select

        options = select_element.find_elements(By.TAG_NAME, "option")

        pattern = re.compile(r'_([^_]+)')
        selectores = [pattern.search(option.text).group(1) for option in options if pattern.search(option.text)]

        eliminar_serv = []
        avisounavez = False    
        avisounavez2 = False   
        elementos_tr = driver.find_elements(By.CSS_SELECTOR, 'tr[id^="W0046GridContainerRow_"]')
        n_elementos = len(elementos_tr)
        
        list_servis = cadenaserv.split("\n")
        list_escritura = cadena_escritura.split("\n")
        list_area = cadena_area.split("\n")
        
        
        for servis, escri_servi, area_servi in zip(list_servis,list_escritura,list_area):
            textito = ''
            dic_serv = {}
            for j in selectores:
                if j.upper() in servis:
                    dic_serv[j] = False

            
            n_escritura, fecha = escri_servi.lower().split(" del ")
            fecha = datetime.strptime(fecha, "%Y-%m-%d")
            date_doc = fecha.strftime('%d/%m/%Y')
            #date_doc = date_doc.strftime("%d/%m/%Y")
            
            primero = False
            for arreglo in dic_serv:
                if dic_serv[arreglo] == False:
                    if not primero:
                        primero = True
                        continue
                    else:
                        textito = f"CON SERV. DE {arreglo} "
            textito = textito + f"ESTABLECIDA(S) POR ESCRITURA {n_escritura} DE {date_doc}"
            if area_servi != 'NO':
                textito = textito + f" CON UN AREA DE {area_servi} M2"
            
            # n_ser_encon = 1000000
            if not del_every_serv:
                add_serv = dic_serv.copy()
                for k in range(n_elementos):                
                    try:
                        elemento_span = driver.find_element(By.ID,f"span_W0046ACTSERVTIPOID_000{k+1}")
                        valor_seleccionado = elemento_span.text
                        
                        elemento_span2 = driver.find_element(By.ID,f"span_W0046ACTSERVIDUMBREOBS_000{k+1}")
                        # Obtener el texto dentro del elemento
                        observacion = elemento_span2.text
                    except:
                        valor_seleccionado = ""
                        observacion = ""
                    
                    for a in add_serv:
                        add_serv[a] = False
                    
                    for index,i in enumerate(dic_serv.keys()):
                        if i in valor_seleccionado and str(n_escritura) in observacion and dic_serv[i] != True:
                            
                            dic_serv[i] = True
                            add_serv[i] = True
                        
                        # elif dic_serv[i]:
                        #     eliminar_serv.append(k)
                        elif all(valor == False for valor in add_serv.values()) and index == len(dic_serv) - 1:
                            eliminar_serv.append(k)
                            cambiosQC.value = cambiosQC.value + '\nServidumbre modificado'
            elif not avisounavez:
                for k in range(n_elementos):  
                    for index,i in enumerate(dic_serv.keys()):
                            eliminar_serv.append(k)
                            cambiosQC.value = cambiosQC.value + '\nServidumbre modificado'
                avisounavez = True

            if not avisounavez2:
                sustractor = 0
                for clave in eliminar_serv:

                    try:
                        driver.find_element(By.CSS_SELECTOR, f"#span_W0046ACTSERVTIPOID_000{clave+1 - sustractor}")
                        time.sleep(1)
                        driver.find_element(By.CSS_SELECTOR, f"#W0046vDELETE_000{clave+1 - sustractor}").click()
                        time.sleep(1)
                        wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, "body > div.gx-mask.gx-unmask")))
                    except:
                        print ("")
                    sustractor = sustractor + 1
                avisounavez2 = True
            
            arreglo = list(dic_serv.keys())[0]
            time.sleep(3)
            select_element.send_keys("Servidumbre_"+arreglo)
            time.sleep(1)
            driver.find_element(By.CSS_SELECTOR,"#W0046vACTSERVIDUMBREOBS").send_keys(textito)
            time.sleep(1)
            wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "#W0046vACTSERVTIPOID")))
            driver.find_element(By.CSS_SELECTOR, "#W0046ENTER").click()
            time.sleep(2)
            wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, "body > div.gx-mask.gx-unmask")))
            time.sleep(1)
            select_element = driver.find_element(By.CSS_SELECTOR,"#W0046vACTSERVTIPOID")

def crear_interes(pn,sn,pa,sa,genre,conteo_ced,cedula,ced_intered_found = ''):
    # Abre el archivo de Excel R1
    archivo_excel = openpyxl.load_workbook('R1.xlsx')
    # Selecciona la hoja en la que deseas buscar
    hoja = archivo_excel['Hoja1']
    archivo_excel.close()
    
    wait = WebDriverWait(driver, 10)
    panel = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "#Tab_TAB1Containerpanel6")))
    panel.click()
    time.sleep(1.5)
    paneol = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "#W0054CREARINTERESADO")))
    time.sleep(1.5)
    wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, "body > div.gx-mask.gx-unmask")))
    paneol.click()
    time.sleep(1.5)


    wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "#vSOLICITANTE_ID")))
    persona = driver.find_element(By.CSS_SELECTOR,"#vSOLICITANTE_ID")
    time.sleep(2)
    if genre != "N":                                        #N: NIT, H: hombre, M: mujer
        posicion_id = 0                                     #posicion del ciclo en el while
        fix_name = False                              #Arreglar nombre
        fix_ced = False
        tipo_id_selecc_str = ""
        cedula_selecc = ""
        editar_selecc = ""
        sn_sa_blank = False                                 #Para saber cuando sn y sa no tienen nada y volver a pasar
        cambiosn_unavez = False
        same_nombres = 0                                    #cuantos nombres iguales se encontraron
        while True:
            if sn_sa_blank == True and posicion_id == 0 and cambiosn_unavez:
                cambiosn_unavez = False
                time.sleep(.5)
                ay = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "#LIMPIARCAMPOS")))
                ay.click()
                time.sleep(1.5)
                sn = pa
                pa = ''
            
            persona.send_keys("PERSONA NATURAL")
            
            vernombre = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "#vACTMIENINTEPNOMBRE")))
            vernombre.clear()
            vernombre.send_keys(pn)
            pepo = driver.find_element(By.CSS_SELECTOR, "#vACTMIENINTEPAPELLIDO")
            pepo.clear()
            pepo.send_keys(pa)
            
            if ced_intered_found != '':
                time.sleep(1)
                driver.find_element(By.CSS_SELECTOR, "#vACTMIENINTENUMDOC").clear()
                driver.find_element(By.CSS_SELECTOR, "#vACTMIENINTENUMDOC").send_keys(ced_intered_found)
                wait.until(EC.text_to_be_present_in_element_value((By.CSS_SELECTOR, "#vACTMIENINTENUMDOC"), str(cedula)))  
            
            tipo_id = ["Cedula Ciudadanía","Tarjeta Identidad","Secuencial"]
            
            while True:
                
                try: 
                    driver.find_element(By.CSS_SELECTOR, "#vTIPOIDENTIFICACION_ID").send_keys(tipo_id[posicion_id])
                    time.sleep(2)
                    # if posicion_id == 2:
                    #     print ("hola")
                    if posicion_id == 0:
                        driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
                        time.sleep(.5)
                        driver.find_element(By.CSS_SELECTOR, "#GRIDPAGINATIONBARContainer_DVPaginationBar > div.PaginationBarCaption.dropdown > div > button").click()
                        driver.find_element(By.XPATH, "//ul/li/a/span[contains(text(), 'All rows')]").click()
                        time.sleep(.5)
                        driver.execute_script("window.scrollTo(0,0)")
                        
                    time.sleep(1)
                    wait = WebDriverWait(driver, 10)
                    wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, "body > div.gx-mask.gx-unmask")))
                    
                    # Busca todos los elementos <tr> en la página
                    elementos_tr = driver.find_elements(By.TAG_NAME, "tr")
                    # print (elementos_tr)
                    #Itera a través de los elementos <tr>
                    cuantos = -1
                    same_nombres_ciclo = 0
                    
                    
                    for elemento in elementos_tr:
                        # Busca el elemento td en la tercera posición (índice 2)
                        nombre_elemento = elemento.find_elements(By.CSS_SELECTOR, "td:nth-child(3) span")
                        cedula_elemento = elemento.find_elements(By.CSS_SELECTOR, "td:nth-child(5) span")
                        
                        if cedula_elemento:
                            nombre = quitar_acentos(nombre_elemento[0].text)
                            
                            cuantos += 1
                            print(nombre)   #Saber qué nombres se enlistan
                            if (nombre == pn+sn+pa+sa) or ((nombre == pn+pa+sa) and (not sn_sa_blank)) or cedula_elemento[0].text == str(cedula):
                                if (cedula_elemento[0].text == str(cedula)) and (nombre != pn+sn+pa+sa):
                                    fix_name = True
                                tipo_id_selecc_str = tipo_id[posicion_id]                                               
                                same_nombres +=1
                                same_nombres_ciclo += 1
                                #REVISAR SI SE DEBE PONER EL same_nombres > 1: ACA
                                cedula_selecc = elemento.find_elements(By.CSS_SELECTOR, "td:nth-child(5) span")
                                cedula_selecc = cedula_selecc[0].text
                                genero_selecc = elemento.find_elements(By.CSS_SELECTOR, "td:nth-child(8) span")
                                genero_selecc = genero_selecc[0].text
                                editar_selecc = elemento.get_attribute("data-gxrow")
                                # Itera a través de las filas de la columna H
                                if cedula_selecc != str(cedula):
                                    for fila in hoja.iter_rows(min_col=8, max_col=11, values_only=True):                     #COMPARAR CEDULA CON ALGUNA EN R1
                                        # Verifica si la cadena de texto que estás buscando se encuentra en la celda
                                        sin_acento = quitar_acentos(fila[0])
                                        nombres_posibles = [pn + sn + pa + sa,
                                                            " " + pn + sn + pa + sa
                                                            ]
                                        if not sn_sa_blank:
                                            nombres_posibles.append( pn + pa + sa)
                                            nombres_posibles.append(" " + pn + pa + sa)
                                        if any(nombre in sin_acento for nombre in nombres_posibles):
                                            if cedula_selecc != (fila[3] and '0') or '-' in cedula_selecc:
                                                fix_ced = True 
                                                cedula = fila[3]
                                                cambiosQC.value = cambiosQC.value + f'\nCedula de {nombres_posibles[0]} de R1'
                                        else:
                                            fix_ced = True
                                                            
                    if same_nombres_ciclo == 0:
                        raise Exception()
                    elif same_nombres > 1:
                        print ("MISMO NOMBRE ENCONTRADO DOS VECES, REVISAR: ", nombre_elemento)
                        break
                    else:
                        break
                except:
                    popi = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "#vSOLICITANTE_ID")))
                    # Crea un objeto Select para interactuar con el ComboBox
                    select = Select(popi)
                    # Obtén el valor actual seleccionado en el ComboBox
                    valor_actual = select.first_selected_option.text
                    if valor_actual == "Seleccione...":
                        driver.refresh()
                        break
                    
                    posicion_id += 1
                    if posicion_id > 2:
                        if sn == '' and sa == '' and same_nombres == 0 and not sn_sa_blank:
                            cambiosn_unavez = True
                            posicion_id = -1
                            sn_sa_blank = True
                        break
            
            if posicion_id >= 2:
                break
            posicion_id += 1        
                            #print(nombre)   #Saber qué nombres se enlistan
                    #print (cuantos)         #Saber cuantos nombres hay listados
        if cedula == None:
            cedula = ""
        
        if same_nombres == 0 and cedula == "":
            fix_ced = True
        
        cedula_cero = False
        if cedula == "" and fix_ced:                #para que si no hay cedula, se establezca una cedula 9 98 987 etc
                cedula_cero = True
                cedula = "9" * 2 + "9" * (conteo_ced - 1)
                    
        if genre == "H":
            genre = "MASCULINO"
        elif genre == "M":
            genre = "FEMENINO"          
        
        
        
        if same_nombres != 0:
            wait = WebDriverWait(driver, 10)
            time.sleep(.4)
            driver.find_element(By.CSS_SELECTOR, "#vTIPOIDENTIFICACION_ID").send_keys(tipo_id_selecc_str)
            time.sleep(1.5)
           
            editar_inter = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "#vUPDATE_"+editar_selecc))) # Ajusta el selector según tu página.
            wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, "body > div.gx-mask.gx-unmask"))) 
            edito = driver.find_element(By.CSS_SELECTOR, "#vUPDATE_"+editar_selecc)
            
            driver.execute_script("arguments[0].scrollIntoView();", edito)   
            time.sleep(.5)     
    
            encontrado = False
            conta = 0
            vayase = False
            while not encontrado:
                try:
                    # Intenta encontrar el elemento
                    editar_inter.click()
                    encontrado = True
                except:
                    
                    # Si no se encuentra el elemento, desplaza el scroll hacia arriba
                    driver.execute_script("window.scrollBy(0, -100);")
                    # Espera un momento para que la página se cargue y se actualice
                    time.sleep(.7)
                    if conta > 8:
                        vayase = True
                        break
                    conta += 1
            if vayase == True:
                raise Exception("ERROR: Se dañó en el ciclo del while de interesados") 
            time.sleep(.5)  
            
            driver.execute_script("window.scrollTo(0,0)")
            time.sleep(1)
            
            if fix_ced:
                driver.find_element(By.CSS_SELECTOR, "#vACTMIENINTENUMDOC").clear()
                driver.find_element(By.CSS_SELECTOR, "#vACTMIENINTENUMDOC").send_keys(cedula)
                wait.until(EC.text_to_be_present_in_element_value((By.CSS_SELECTOR, "#vACTMIENINTENUMDOC"), str(cedula)))  
                time.sleep(2)
            #wait.until(EC.text_to_be_present_in_element_value((By.CSS_SELECTOR, "#vACTMIENINTENUMDOC"), cedula_selecc))  
            
            driver.find_element(By.CSS_SELECTOR, "#vTIPOIDENTIFICACION_ID").send_keys(tipo_id[0])
            time.sleep(1.5)
            
            driver.find_element(By.CSS_SELECTOR, "#vACTGENEROID").send_keys(genre)
            time.sleep(2)
            driver.find_element(By.CSS_SELECTOR, "#vACTGRUPOETNICOID").send_keys("Ninguno")
            time.sleep(2)
            
            if fix_name:
                driver.find_element(By.CSS_SELECTOR, "#vACTMIENINTESNOMBRE").clear()
                driver.find_element(By.CSS_SELECTOR, "#vACTMIENINTESNOMBRE").send_keys(sn)
                time.sleep(1)
                driver.find_element(By.CSS_SELECTOR, "#vACTMIENINTESAPELLIDO").clear()
                driver.find_element(By.CSS_SELECTOR, "#vACTMIENINTESAPELLIDO").send_keys(sa)
                time.sleep(2)
            
            if sn_sa_blank:            
                pa = sn
                sn = ''
                driver.find_element(By.CSS_SELECTOR, "#vACTMIENINTEPAPELLIDO").clear()
                driver.find_element(By.CSS_SELECTOR, "#vACTMIENINTEPAPELLIDO").send_keys(pa)
                time.sleep(1)
                driver.find_element(By.CSS_SELECTOR, "#vACTMIENINTESNOMBRE").clear()

            # if fix_gen:
            #     driver.find_element(By.CSS_SELECTOR, "#vACTGENEROID").send_keys(genre)
            # if tipo_id_selecc:
            #     driver.find_element(By.CSS_SELECTOR, "#vTIPOIDENTIFICACION_ID").send_keys(tipo_id[0])
        else:
            time.sleep(1.5)
            driver.find_element(By.CSS_SELECTOR, "#vACTMIENINTENUMDOC").send_keys(cedula)
            time.sleep(1.5)
            driver.find_element(By.CSS_SELECTOR, "#vACTGENEROID").send_keys(genre)
            time.sleep(2)
            driver.find_element(By.CSS_SELECTOR, "#vTIPOIDENTIFICACION_ID").send_keys(tipo_id[0])
            time.sleep(1.5)
            driver.find_element(By.CSS_SELECTOR, "#vACTGRUPOETNICOID").send_keys("Ninguno")
            time.sleep(1)
            driver.find_element(By.CSS_SELECTOR, "#vACTMIENINTEPNOMBRE").clear()
            driver.find_element(By.CSS_SELECTOR, "#vACTMIENINTEPNOMBRE").send_keys(pn)
            time.sleep(1)
            driver.find_element(By.CSS_SELECTOR, "#vACTMIENINTEPAPELLIDO").clear()
            driver.find_element(By.CSS_SELECTOR, "#vACTMIENINTEPAPELLIDO").send_keys(pa)
            time.sleep(1)
            driver.find_element(By.CSS_SELECTOR, "#vACTMIENINTESNOMBRE").send_keys(sn)
            time.sleep(1)
            driver.find_element(By.CSS_SELECTOR, "#vACTMIENINTESAPELLIDO").send_keys(sa)
        
    else:
        time.sleep(.5)
        ay = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "#LIMPIARCAMPOS")))
        ay.click()
        time.sleep(1)
        cedula_cero = False
        persona.send_keys("PERSONA JURIDICA")
        time.sleep(1)
        driver.find_element(By.CSS_SELECTOR, "#vTIPOIDENTIFICACION_ID").send_keys("Numero de Identificación Tributaria")
        time.sleep(1)
        elemento = driver.find_element(By.CSS_SELECTOR, "#GRIDPAGINATIONBARContainer_DVPaginationBar > div")
        if "PaginationBarEmptyGrid" not in elemento.text:
            if cedula == "":                #para que si no hay cedula, se establezca una cedula 9 98 987 etc
                cedula_cero = True
                cedula = "9" * 2 + "9" * (conteo_ced - 1)
            time.sleep(1)
            driver.find_element(By.CSS_SELECTOR, "#vACTMIENINTERAZSOC").send_keys(pn+sn+pa+sa)
            time.sleep(1)
            driver.find_element(By.CSS_SELECTOR, "#vACTMIENINTEDIGIVERI").send_keys("0")
        cedula = cedula.split("-")[0]
        driver.find_element(By.CSS_SELECTOR, "#vACTMIENINTENUMDOC").send_keys(cedula)
        time.sleep(2)
        driver.find_element(By.CSS_SELECTOR, "#GUARDAR").click()
        time.sleep(2)
        wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, "body > div.gx-mask.gx-unmask")))
  
    time.sleep(2)
    driver.find_element(By.CSS_SELECTOR, "#GUARDAR").click()
    time.sleep(2)
    wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, "body > div.gx-mask.gx-unmask")))
    papa = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "#CANCEL")))
    papa.click()
    time.sleep(1.5)
    return cedula,cedula_cero
        

def ajustar_numero(cadena):
    # Verificar si la cadena tiene decimales
    cadena = str(cadena)
    if cadena == '':
        cadena = '0'
    if '.' in cadena:
        partes = cadena.split('.')
        parte_decimal = partes[1]
        
        # Verificar si hay más de dos dígitos decimales
        if len(parte_decimal) > 2:
            # Si hay más de dos dígitos, verificar si los siguientes son '0'
            if parte_decimal[2:].isdigit() and int(parte_decimal[2:]) == 0:
                # Eliminar los dígitos decimales extras
                cadena = partes[0] + ',' + parte_decimal[:2]

        cadena = cadena.replace('.', ',')
        if len(parte_decimal) == 1:
            cadena += '0'
    else:
        # Si no tiene decimales, agregar una coma y dos ceros al final
        cadena += ',00'
    return cadena
        
def llenar_predio(hojainfo,mat_matriz,direccion,areaTerr,coef):
    wait = WebDriverWait(driver, 10)
    matricula = driver.find_element(By.CSS_SELECTOR, "#Tab_TAB1Containerpanel1")
    matricula.click()
    time.sleep(1)
    
    informacion_combobox = {
        '#W0014vACTPREDIOMATRICULAMATR': mat_matriz,
        '#W0014vACTPREDIODIREPRIP': direccion,
        '#W0014vACTPREDIOARETERR': areaTerr,
    }     
    names_combo =["Matricula Matriz","Direccion","Area de terreno"]
    for (combobox_id, opcion), names in zip(informacion_combobox.items(), names_combo):
        element = driver.find_element(By.CSS_SELECTOR, combobox_id)

        valor = element.get_attribute("value")
        
        if combobox_id == '#W0014vACTPREDIOARETERR':
            opcion = ajustar_numero(opcion)
        
        if combobox_id == '#W0014vACTPREDIODIREPRIP' and valor.strip() == '' and opcion == '':
            element.send_keys("SIN DIRECCION")
            cambiosQC.value = cambiosQC.value + f'\nSin direccion en Multicarto ni en VUR'
        
        if valor.strip() and opcion == '' and combobox_id == '#W0014vACTPREDIODIREPRIP':
            print ("")
        elif valor.strip() == opcion.strip():
            #cambiosQC.value = cambiosQC.value + f'\n{names} ta bien'
            print ("")
        else:  
            element.clear()
            time.sleep(.5)
            element.send_keys(opcion)
            cambiosQC.value = cambiosQC.value + f'\n{names} arreglada'
        # else:
        #     element.clear()
        #     time.sleep(.5)
        #     element.send_keys(opcion)
    
    cambioestado = driver.find_element(By.CSS_SELECTOR, "#W0014vACTPREDIOESTADO")
    time.sleep(1)
    if poneraprobado:
        cambioestado.send_keys("Aprobado")
    else:
        cambioestado.send_keys("Realizado")
    
    combodestino = driver.find_element(By.CSS_SELECTOR, "#W0014vDESTINACIONID")
    
    # Utilizando Select para manejar elementos select
    select = Select(combodestino)
    texto_actual = select.first_selected_option.text.strip()

    if texto_actual == 'Seleccione...':
        select.select_by_visible_text('LOTE URBANIZADO NO CONSTRUIDO')  # Seleccionar por texto visible
    else:
        print("Destino no cambiado")
    
    wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, "body > div.gx-mask.gx-unmask")))
    checkbox_element = driver.find_element(By.CSS_SELECTOR, "#W0014vACT_PREDIOAREAREGISTRAL")      # BOTON DE QUE NO HAY AREA CATASTRAL
    if not checkbox_element.is_selected():
        checkbox_element.click()
    # Encuentra el elemento por su ID
    elemento = driver.find_element(By.CSS_SELECTOR, "#W0014vACTPREDIOARECONST")

    # Añade un nuevo dato a la celda a la derecha de "Dirección Corregida"
    nueva_celda = hojainfo.cell(row=fila_a_extraer, column=columna_max + 1)
    nueva_celda.value = elemento.get_attribute("value")  # Reemplaza con el valor que desees
    coef_exist = driver.find_element(By.CSS_SELECTOR, "#span_W0014vACTPREDIOCOEPRE")
    coef_exist2 = driver.find_element(By.CSS_SELECTOR, "#W0014vACTPREDIOCOEPRE")
    text_coef2 = coef_exist2.get_attribute("value")
    text_coef = coef_exist.text
    driver.find_element(By.CSS_SELECTOR, "#W0014GUARDAR").click()
    
    if coef != '' and ('0,00' in text_coef or '0,00' in text_coef2):
        oe = driver.find_element(By.CSS_SELECTOR, "#W0014vACTPREDIOCOEPRE")
        time.sleep(1)
        oe.send_keys(coef)
        driver.find_element(By.CSS_SELECTOR, "#W0014GUARDAR").click()
    
    time.sleep(1)
    wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, "body > div.gx-mask.gx-unmask")))



def crear_fuente(fuente,date_doc,ente,date_reg,n_fuente="SN"):

    matricula = driver.find_element(By.CSS_SELECTOR, "#Tab_TAB1Containerpanel3")
    matricula.click()
    time.sleep(1)
    # date_doc = date_doc.replace("-","/")
    # date_reg = date_reg.replace("-","/")
    # date_doc = datetime.strptime(date_doc,"%d/%m/%Y")
    
    # Supongamos que date_doc puede ser una cadena o un objeto datetime
    if isinstance(date_doc, str):
        date_doc = datetime.strptime(date_doc, "%d-%m-%Y")
    # Ahora puedes realizar las operaciones de formato si date_doc es de tipo datetime
    date_doc = date_doc.strftime("%d/%m/%Y")
    
    if isinstance(date_reg, str):
        date_reg = datetime.strptime(date_reg, "%d-%m-%Y")
    # Ahora puedes realizar las operaciones de formato si date_doc es de tipo datetime
    date_reg = date_reg.strftime("%d/%m/%Y")
    
    # Define un diccionario con los selectores CSS y sus valores
    campos = {
        "#W0030vACTFUENTEADMINTIPOID": fuente,
        "#W0030vACTPRINCIPALDOCTIPOID": "Documento",
        "#W0030vACTNUMFUENTE": n_fuente,
        "#W0030vACTENTEEMISOR": ente
    }
    if fuente == "RESOLUCION":
        campos['#W0030vACTFUENTEADMINTIPOID'] = "Acto"
    
    # Llena los campos utilizando un bucle
    for selector, valor in campos.items():
        campo = driver.find_element(By.CSS_SELECTOR, selector)
        if valor == n_fuente:
            campo.clear()    
            time.sleep(.5)
        campo.send_keys(valor)
        time.sleep(.5)

    # driver.find_element(By.CSS_SELECTOR, "#W0030vACTFUEADMFECHDOC").send_keys(date_doc.strftime("%d/%m/%Y"))
    # driver.find_element(By.CSS_SELECTOR, "#W0030vACTFUEADMFECHREG").send_keys(date_reg.strftime("%d/%m/%Y"))
    # Completa las fechas utilizando driver.execute_script
    driver.execute_script("arguments[0].value = arguments[1];", driver.find_element(By.CSS_SELECTOR, "#W0030vACTFUEADMFECHDOC"), date_doc)
    driver.execute_script("arguments[0].value = arguments[1];", driver.find_element(By.CSS_SELECTOR, "#W0030vACTFUEADMFECHREG"), date_reg)

    matricula = driver.find_element(By.CSS_SELECTOR, "#W0030ENTER")
    matricula.send_keys(Keys.PAGE_DOWN)
    matricula.click()
    
    time.sleep(1)
    wait = WebDriverWait(driver, 15)
    wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, "body > div.ui-pnotify.stack-topright")))
    
def subir_documentos(ruta,cadena_folio):
    driver.find_element(By.CSS_SELECTOR, "#Tab_TAB1Containerpanel4").click()
    combo = driver.find_element(By.CSS_SELECTOR, "#W0038vACTFUENTADMINID")
    select = Select(combo)
    select.select_by_index(1)
    
    file_input = driver.find_element(By.CSS_SELECTOR,"#W0038FILEUPLOAD1Container > div > div.row.fileupload-buttonbar > div.col-lg-7 > span.btn.btn-success.fileinput-button > input[type=file]")  # Reemplaza "fileInput" con el ID de tu elemento

    # Usa send_keys() para proporcionar la ruta completa del archivo a cargar
    time.sleep(1.5)
    file_input.send_keys(ruta+cadena_folio+" J.pdf")
    #time.sleep(1)
    file_input = driver.find_element(By.CSS_SELECTOR,"#W0038FILEUPLOAD1Container > div > div.row.fileupload-buttonbar > div.col-lg-7 > span.btn.btn-success.fileinput-button > input[type=file]")  # Reemplaza "fileInput" con el ID de tu elemento
    time.sleep(1.5)
    file_input.send_keys(ruta+cadena_folio+" B.pdf")
    driver.find_element(By.CSS_SELECTOR, "#W0038FILEUPLOAD1Container > div > div.row.fileupload-buttonbar > div.col-lg-7 > button.btn.btn-primary.start").click()
    wait = WebDriverWait(driver, 10)
    time.sleep(3)
    wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, "#W0038FILEUPLOAD1Container > div > table > tbody > tr:nth-child(1) > td:nth-child(4) > button.btn.btn-primary.start")))
    
    driver.find_element(By.CSS_SELECTOR, "#W0038CONFIRMAR").click()
    time.sleep(1.5)
    wait = WebDriverWait(driver, 10)
    wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, "body > div.ui-pnotify.stack-topright")))
 
def derechos(texto):
    wait = WebDriverWait(driver, 10)
    pato = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "#Tab_TAB1Containerpanel8")))
    pato.click()

    selectores = ["#W0070vACTFUENTADMINID", "#W0070vACTGRUPINTEREID", "#W0070vACTDERECHOTIPOID"]

    for selector in selectores:
        combo = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, selector)))
        combo = driver.find_element(By.CSS_SELECTOR, selector)
        select = Select(combo)
        wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, f"{selector} option:nth-child({1 + 1})")))
        if selector == "#W0070vACTDERECHOTIPOID" and ("COMPRAVENTA DERE" in texto or "FALSA TRADICION" in texto or "ADJUDICACION" in texto):
            combo.send_keys("POSESION")
        time.sleep(3)
        select.select_by_index(1)

    driver.find_element(By.CSS_SELECTOR, "#W0070GUARDAR").click()
    wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, "body > div.gx-mask.gx-unmask")))
    
    
    
 
def set_zero_intereados(list_identificacion, lis_cedulas,list_booleanos):
    wait = WebDriverWait(driver, 10)
    time.sleep(1.5)
    oe = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "#Tab_TAB1Containerpanel6")))
    oe.click()
    eo = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "#W0054CREARINTERESADO")))
    eo.click()    
    
    for identificadores, ceduloides, booleanos in zip(list_identificacion, lis_cedulas,list_booleanos):
        if booleanos == True:
            error_count = 0
            ocurre_error = False
            while True:
                try:
                    wait.until(lambda driver: len(driver.find_elements(By.CSS_SELECTOR, ".ui-pnotify.stack-topright[style*='display: block']")) <= 1)
                    time.sleep(.5)
                    ay = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "#LIMPIARCAMPOS")))
                    ay.click()
                    try:
                        wait = WebDriverWait(driver, 10)
                        wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, "body > div.gx-mask.gx-unmask")))
                    except:
                        driver.refresh()    
                    time.sleep(1.5)
                    if identificadores != "N":                #N: NIT, H: hombre, M: mujer
                        wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "#vSOLICITANTE_ID"))).send_keys("PERSONA NATURAL")
                        wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "#vACTMIENINTEPNOMBRE")))
                        time.sleep(1)
                        driver.find_element(By.CSS_SELECTOR, "#vTIPOIDENTIFICACION_ID").send_keys("Cedula Ciudadanía")
                    else:
                        time.sleep(1.5)
                        wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "#vSOLICITANTE_ID"))).send_keys("PERSONA JURIDICA")
                        wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "#vACTMIENINTERAZSOC")))
                        time.sleep(1.5)
                        driver.find_element(By.CSS_SELECTOR, "#vTIPOIDENTIFICACION_ID").send_keys("Numero de Identificación Tributaria")
                    
                    driver.find_element(By.CSS_SELECTOR, "#vACTMIENINTENUMDOC").send_keys(ceduloides)
                    wait.until(EC.text_to_be_present_in_element_value((By.CSS_SELECTOR, "#vACTMIENINTENUMDOC"), ceduloides))  
                    
                    time.sleep(1.5)
                    paco = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "#GRIDPAGINATIONBARContainer_DVPaginationBar > div.PaginationBarCaption.dropdown > div > button")))
                    paco.click()
                    driver.find_element(By.XPATH, "//ul/li/a/span[contains(text(), 'All rows')]").click()
                    
                    wait = WebDriverWait(driver, 10)
                    wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, "body > div.gx-mask.gx-unmask")))
                    time.sleep(5)
                    # Busca todos los elementos <tr> en la página
                    elementos_tr = driver.find_elements(By.TAG_NAME, "tr")

                    counta = 0
                    cc_repetido = False
                    editar_selecc = [] 
                    genero_selec = []
                    for elemento in elementos_tr:
                        # Busca el elemento td en la tercera posición (índice 2)
                        nombre_elemento = elemento.find_elements(By.XPATH, ".//td[@data-colindex='4']//span")

                        if nombre_elemento:
                            nombre = nombre_elemento[0].text
                            gen = elemento.find_elements(By.XPATH, ".//td[@data-colindex='7']//span")
                            if nombre == ceduloides:
                                editar_selecc.append(elemento.get_attribute("data-gxrow"))
                                genero_selec.append(gen[0].text)
                                if counta >= 1:
                                    counta += 1
                                    break
                                counta += 1

                    for be in range(counta):
                        # time.sleep(1.5)
                        # pepo = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "#vDELETE_"+editar_selecc[be]))) # Ajusta el selector según tu página.
                        # wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, "body > div.gx-mask.gx-unmask")))
                        # time.sleep(1.5)
                        # pepo.click()
                        
                        pepo = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "#vUPDATE_"+editar_selecc[be]))) # Ajusta el selector según tu página.
                        #pepo = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "#vUPDATE_0001"))) # Ajusta el selector según tu página.
                        wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, "body > div.gx-mask.gx-unmask")))
                        time.sleep(1.5)
                        pepo.click()
                        time.sleep(1.5)
                        if identificadores != "N":                #N: NIT, H: hombre, M: mujer
                            driver.find_element(By.CSS_SELECTOR, "#vACTGRUPOETNICOID").send_keys("Ninguno")
                            driver.find_element(By.CSS_SELECTOR, "#vACTGENEROID").send_keys(genero_selec[be])
                            
                        driver.find_element(By.CSS_SELECTOR, "#vACTMIENINTENUMDOC").clear()
                        time.sleep(.5)
                        driver.find_element(By.CSS_SELECTOR, "#vACTMIENINTENUMDOC").send_keys("0")
                        wait.until(EC.text_to_be_present_in_element_value((By.CSS_SELECTOR, "#vACTMIENINTENUMDOC"), "0"))  
                        time.sleep(1)
                        driver.find_element(By.CSS_SELECTOR, "#GUARDAR").click()
                        time.sleep(1)
                        driver.find_element(By.CSS_SELECTOR, "#GUARDAR").click()
                        wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, "body > div.gx-mask.gx-unmask")))
                        time.sleep(2)               
                    break
                except:
                    driver.refresh()
                    if error_count > 6:
                        ocurre_error = True
                        break
                    error_count += 1
            if ocurre_error:
                break
    if ocurre_error:
        raise Exception("ERROR: Se dañó en def_zero_empleados")     
    driver.find_element(By.CSS_SELECTOR, "#CANCEL ").click()
    wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, "body > div.gx-mask.gx-unmask")))
    time.sleep(1.5)

        

def borrar_archivos(archivolio,cantidad):

    wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, "body > div.gx-mask.gx-unmask")))
    driver.find_element(By.CSS_SELECTOR, "#Tab_TAB1Containerpanel4").click()
    combo = driver.find_element(By.CSS_SELECTOR, "#W0038vACTFUENTADMINID")
    select = Select(combo)
    select.select_by_index(1) 

    
    try:
        wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, "body > div.gx-mask.gx-unmask")))
        wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, 'img[id^="W0038vBORRAR_000"]')))
        borrar = driver.find_element(By.CSS_SELECTOR,'img[id^="W0038vBORRAR_0001"]')             #BORRAR DOCUMENTOS
        borrar.click()
        time.sleep(1.2)
        
        if cantidad == 2:
            wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, "body > div.gx-mask.gx-unmask")))
            bobo= wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, 'img[id^="W0038vBORRAR_000"]')))
            bobo.click()
    except: 
        print ("Archivos eliminados")
    try:
        wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, "#gxHTMLWrpW0038 > div.gx-mask")))
    except:
        driver.refresh()
        
            
def buscar_inter_malo(driver, datosjuridicos, sininteres, unaveznomas,n_ciclo,max_ciclo):       #n_ciclo: para saber en qué # de cedula va, 
    buscar_malos = [False, False, True, False]                                                  #max_ciclo: cedulas totales a revisar
    cedsinborrar = ""
    contador_findnames = 0
    bool_avisar = False                      #avisa que el # de interesados es mayor en MT que en el excel
    cadena1_interesado = ""                 #Para ver los datos del interesado
    if datosjuridicos["Género"] == 'N':
        nombre_completo = datosjuridicos["Primer Nombre"].strip()
    else:
        nombre_completo = datosjuridicos["Primer Nombre"] + datosjuridicos["Segundo Nombre"] + datosjuridicos["Primer Apellido"] + datosjuridicos["Segundo Apellido"]
    nombre_sin_segundo_nombre = datosjuridicos["Primer Nombre"] + datosjuridicos["Primer Apellido"] + datosjuridicos["Segundo Apellido"]

    errordir={}
    if sininteres and unaveznomas:
        wait = WebDriverWait(driver, 10)
        wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, "body > div.gx-mask.gx-unmask")))
        time.sleep(1.5)
        elementos_tr = driver.find_elements(By.TAG_NAME, "tr")
        for elemento in elementos_tr:
            # driver.execute_script("window.scrollTo(0,0)")
            # time.sleep(.5)
            # encontrar = elemento.find_elements(By.CSS_SELECTOR, "td:nth-child(5) span")
            # search_element_click(encontrar[0])
            
            if datosjuridicos["Género"] == 'N':
                buscar_nombre = elemento.find_elements(By.CSS_SELECTOR, "td:nth-child(9) span")
            else:
                buscar_nombre = elemento.find_elements(By.CSS_SELECTOR, "td:nth-child(7) span")

            if buscar_nombre:
                buscar_nombre = quitar_acentos(buscar_nombre[0].text)
                print(buscar_nombre)
                
                if buscar_nombre != "":
                    contador_findnames += 1
                    
                if  nombre_completo in buscar_nombre or nombre_sin_segundo_nombre in buscar_nombre:
                    buscar_malos[2] = False
                    buscar_tipo = elemento.find_elements(By.CSS_SELECTOR, "td:nth-child(5) span")
                    buscar_tipo = buscar_tipo[0].text
                    buscar_ced = elemento.find_elements(By.CSS_SELECTOR, "td:nth-child(6) span")
                    buscar_ced = buscar_ced[0].text
                    buscar_gen = elemento.find_elements(By.CSS_SELECTOR, "td:nth-child(10) span")
                    buscar_gen = buscar_gen[0].text
                    cedsinborrar = buscar_ced

                    if (datosjuridicos["Género"] in ("H", "M") and buscar_tipo != "Cedula Ciudadanía") or (datosjuridicos["Género"] == "N" and buscar_tipo != "Numero de Identificación Tributaria"):
                    # Tu lógica si la condición se cumple
                        buscar_malos[0] = True

                    patron = r'\b(?:9|99|99|999|9999)\b'
                    coincidencias = re.findall(patron, buscar_ced)

                    if any(substring in buscar_ced for substring in ("-", '9999')) and datosjuridicos["Género"] in ("H", "M") or coincidencias:
                        buscar_malos[1] = True

                    if datosjuridicos["Género"] in ("H", "M") and buscar_gen == "SIN_DETERMINAR":
                        buscar_malos[3] = True
                    
                    errordir = {
                        buscar_tipo:buscar_malos[0],                #Para ver si el tipo de documento no sea secuencial
                        buscar_ced:buscar_malos[1],                 #Buscar que el cc este bien           
                        buscar_gen: buscar_malos[3]                #Mirar que el genero esté bien
                        
                    }
                    
                    if max_ciclo == n_ciclo+1:
                        bool_avisar = True
                    else:
                        break
                
    if contador_findnames > max_ciclo and bool_avisar:
        buscar_malos[2] = True              
    
    if any(buscar_malos):
        cadena1_interesado = nombre_completo + ": "
        for datosantes,bule in errordir.items():
            if bule == True:
                cadena1_interesado = cadena1_interesado + "," + datosantes
    return buscar_malos,cedsinborrar,cadena1_interesado

# Crea un diccionario para mapear los nombres de encabezados a variables
datos = {
    'Folio': None,
    'Coeficiente': None,
    'Matrícula matriz': None,
    'Área de Terreno': None,
    'Dirección Corregida': None,
}

datosjuridicos = {
    'Folio': None,
    'Servidumbre': None,
    'Escr. Serv': None,
    'Area Serv': None,
    'Fecha registro': None,
    'Fecha documento': None,
    'Fuente adm.': None,
    'N. Fuente': None,
    'Ente Em.': None,
    'Porcentajes':None,
    'Cédulas': None,
    'Primer Nombre': None,
    'Segundo Nombre': None,
    'Primer Apellido': None,
    'Segundo Apellido': None,
    'Género': None,
    'Texto': None
}
# Obtén los valores de los encabezados en la primera fila
for columna in hoja.iter_cols(min_row=1, max_row=1):
    for celda in columna:
        if celda.value in datos:
            datos[celda.value] = celda.column
            columna_max = celda.column

for columna in hojajuridico.iter_cols(min_row=1, max_row=1):
    for celda in columna:
        if celda.value in datosjuridicos:
            datosjuridicos[celda.value] = celda.column
            columna_max_juridico = celda.column

# Navega a la página web deseada
driver.get("https://www.realidad5.com/realmultipropositobarranca/servlet/com.realmultipropositogam.iniciosesion")


while True:
    try:
        # Localiza el campo de texto de usuario por su selector CSS
        campo_usuario = driver.find_element(By.CSS_SELECTOR, "#vUSERNAME")
        campo_usuario.send_keys("jgonzalez.realtix@gmail.com")

        # Localiza el campo de texto de contraseña por su selector CSS
        campo_contrasena = driver.find_element(By.CSS_SELECTOR, "#vUSERPASSWORD")
        #campo_contrasena.send_keys("u9OPY2tnZs*71")   #SAN LUIS
        campo_contrasena.send_keys("JG1100973580")

        entrar = driver.find_element(By.CSS_SELECTOR, "#BTNENTER")
        entrar.click()
        wait = WebDriverWait(driver, 10)
        # time.sleep(10)
        mas = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "#TEXTBLOCKTITLE_MPAGE")))
        break
    except:
        driver.refresh()

fila_a_extraer = 33  # Reemplaza con el número de fila deseado REVISAR 14 Y 15,16
veces_repetir_folio = 5
datos_titulos = datos.copy() 
datos_titulosjuri = datosjuridicos.copy()
flag = 0
flagaprob = False
modoqcl = True

while hoja.cell(row=fila_a_extraer, column=1).value is not None:
    error = hoja.cell(row=fila_a_extraer, column=columna_max+ 1)
    errorzaso = hoja.cell(row=fila_a_extraer, column=columna_max+ 4)
    cambiosQC = hoja.cell(row=fila_a_extraer, column=columna_max+ 5)
    cambiosQC.value = ""
    if modoqc:
        modoqcl = False
    for a in range(veces_repetir_folio):
        try:
            print ("Llenando la fila # ",fila_a_extraer,"...")
            # Extrae los datos de la fila y almacénalos en el diccionario
            nueva_celda1 = hoja.cell(row=fila_a_extraer, column=columna_max + 1).value
            nueva_celda2 = hoja.cell(row=fila_a_extraer, column=columna_max + 2).value
            nueva_celda3 = hoja.cell(row=fila_a_extraer, column=columna_max + 3).value
            nueva_celda4 = hoja.cell(row=fila_a_extraer, column=columna_max + 4).value
            if (nueva_celda1 is not None and nueva_celda1 != "NO SE LLENÓ FOLIO, REVISAR" and 'Se pro' not in str(nueva_celda4)) or nueva_celda3 == "Unfounded" or nueva_celda2 is not None:
                flag = 1
            else:
                fila = hoja[fila_a_extraer]
                for encabezado, columna in datos_titulos.items():
                    if columna:
                        if fila[columna-1].value == None:
                            datos[encabezado] = ""
                        else:
                            datos[encabezado] = fila[columna-1].value
                
                if flag != 1:    
                    driver.get("https://www.realidad5.com/realmultipropositobarranca/servlet/com.realmultipropositogam.wwfichprediact")
                    wait = WebDriverWait(driver, 5)
                    holi = WebDriverWait(driver, 10)
                    while True:
                        try:
                            mas = holi.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "#ACTIONNEW"))) # Ajusta el selector según tu página.
                            mas.click()
                            break
                        except:
                            driver.refresh()
                wait = WebDriverWait(driver, 3)   
                time.sleep(1)
                wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "#vACTPREDIOESTADO")))
                driver.find_element(By.CSS_SELECTOR, "#vACTPREDIOESTADO").send_keys("Activo")
                matricula = driver.find_element(By.CSS_SELECTOR, "#vACTPREDIOMATRICULA")
                matricula.clear()
                matricula.send_keys(datos['Folio'])
                flag = 0
                time.sleep(2)
                                
                while True:
                    try:
                        
                        matricula = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "#vUPDATE_0001")))
                        matricula.click()
                        flag = 0
                        break
                    except:
                        wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "#vACTPREDIOESTADO")))
                        if flag == 0:
                            driver.find_element(By.CSS_SELECTOR, "#vACTPREDIOESTADO").send_keys("Realizado")
                            flag = 2
                        elif flag == 2:                            
                            driver.find_element(By.CSS_SELECTOR, "#vACTPREDIOESTADO").send_keys("Aprobado")
                            if not modoqc and not poneraprobado:
                                flagaprob = True
                            flag = 1
                        elif flag == 1:
                            flagaprob = False
                            nueva_celda = hoja.cell(row=fila_a_extraer, column=columna_max+3)
                            nueva_celda.value = "Unfounded"
                            break
                    
                            

            if flag != 1:
                time.sleep(5)
                wait = WebDriverWait(driver, 15)
                combobox = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "#W0014vTIPOLOTE_ID"))) # Ajusta el selector según tu página.

                # Obtiene el texto del elemento seleccionado.
                elemento_seleccionado = Select(combobox).first_selected_option
                texto_elemento = elemento_seleccionado.text

                # Define la cadena de caracteres que deseas buscar.
                cadena_busqueda = "Unidad_Predial"
                
                
                # Comprueba si la cadena de búsqueda está presente en el texto.
                if pestañapredio:
                    if cadena_busqueda in texto_elemento and datos["Coeficiente"] == "":
                        nueva_celda = hoja.cell(row=fila_a_extraer, column=columna_max+ 2)
                        nueva_celda.value = "Es PH, y no hay coeficiente"
                    else:
                        if cadena_busqueda not in texto_elemento and datos["Coeficiente"] != "":
                            combobox.send_keys("Unidad_Predial")
                        llenar_predio(hoja,datos['Matrícula matriz'],datos['Dirección Corregida'],datos['Área de Terreno'],datos["Coeficiente"])                                                

                if llenarsolopredio:
                    folio_selec = datos['Folio']
                    fila_juridico = 0
                    fila_cedulas = 1
                    juridico = False
                    datofoliosave = ''
                    for filaj in hojajuridico.iter_rows(min_col=1, max_col=columna_max_juridico, values_only=True):                     #revisar en que posicion esta el folio y la cantidad de interesados
                        # Verifica si la cadena de texto que estás buscando se encuentra en la celda
                        if juridico == False:
                            if str(folio_selec) in str(filaj[1]):
                                juridico = True
                                fila_juridico = fila_cedulas
                        
                        elif (filaj[datos_titulosjuri["Folio"]-1] != datofoliosave) or (filaj[datos_titulosjuri["Primer Apellido"]-1] == '' or filaj[datos_titulosjuri["Primer Apellido"]-1] == None):
                            break
                        fila_cedulas += 1 
                        datofoliosave = filaj[datos_titulosjuri["Folio"]-1]
                        
                    cedulas_revisar = fila_cedulas - fila_juridico
                    fila_juridico_inicial = fila_juridico
                    
                    if juridico == False:
                        aviso_vur = hoja.cell(row=fila_a_extraer, column=columna_max+ 4)
                        aviso_vur.value = "NO HUBO COINCIDENCIA FOLIO BASICO CON EL JURIDICO"
                        print ("NO HUBO COINCIDENCIA FOLIO BASICO CON EL JURIDICO: ",datos['Folio'])
                    else:
                        # Extrae los datos de la fila y almacénalos en el diccionario
                        filaj = hojajuridico[fila_juridico]
                        for encabezado, columna in datos_titulosjuri.items():
                            if columna:
                                if filaj[columna-1].value == None:
                                    datosjuridicos[encabezado] = ""
                                elif isinstance(filaj[columna-1].value, str):
                                    datosjuridicos[encabezado] = filaj[columna-1].value.strip()
                                else:
                                    datosjuridicos[encabezado] = filaj[columna-1].value
                    
                    if juridico:
                        wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, "body > div.gx-mask.gx-unmask")))
                        time.sleep(1.2)
                        matricula = driver.find_element(By.CSS_SELECTOR, "#Tab_TAB1Containerpanel3")
                        matricula.click()
                        time.sleep(1)
                        
                        pregunto_derechos = False
                        holi = False
                        holo = False
                        try:
                            elemento_fuente = driver.find_element(By.CSS_SELECTOR, "#span_W0030ACTNUMFUENTE_0001")
                            holi = True
                        except:
                            print ("No se encontró fuente administrativa, ", datosjuridicos["Folio"])
                            holo = True
                            
                        mod_fuente = False
                        if holi and str(datosjuridicos["N. Fuente"]) != elemento_fuente.text:       #MODIFICAR FUENTE: PRIMERO SE REVISA FOLIO SI FUENTE ES DIFERNTE
                            nueva_celda = hoja.cell(row=fila_a_extraer, column=columna_max+ 3)      #SI ES DIFERENTE PREGUNTAR AL USUARIO POR MEDIO DE LA CELDA MISMA
                            if str(nueva_celda.value) == "1":                                            #QUE DONDE ESTÁ "UNFOUNDED", SI EL USUARIO QUIERE QUE SE MODIFIQUE
                                mod_fuente = True                                                   #DEBE PONER UN 1 EN ESA MISMA CELDA
                                nueva_celda.value = "Fuente adm. Mod."
                                cambiosQC.value = cambiosQC.value + '\nFuente administrativa modificada'
                            else:
                                nueva_celda.value = elemento_fuente.text
                        #a = 1/0
                        
                    if juridico != False and not flagaprob: 
                        if meterleservi == True and datosjuridicos["Servidumbre"] != '':
                            crear_servidumbre(datosjuridicos["Servidumbre"],datosjuridicos["Escr. Serv"],datosjuridicos["Area Serv"])

                        if mod_fuente:

                            driver.find_element(By.CSS_SELECTOR, "#Tab_TAB1Containerpanel3").click()
                            wait = WebDriverWait(driver, 10)
                            borrar_fuente = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "#W0030vUPDATE_0001")))
                            borrar_fuente.click()
                            
                            crear_fuente(datosjuridicos["Fuente adm."],
                                        datosjuridicos["Fecha documento"],
                                        datosjuridicos["Ente Em."],
                                        datosjuridicos["Fecha registro"],
                                        datosjuridicos["N. Fuente"]
                                        )
                        elif holo:
                            crear_fuente(datosjuridicos["Fuente adm."],
                                        datosjuridicos["Fecha documento"],
                                        datosjuridicos["Ente Em."],
                                        datosjuridicos["Fecha registro"],
                                        datosjuridicos["N. Fuente"]
                                        )
                        
                        if mod_fuente:
                            modoqcl = True
                        
                        time.sleep(1)
                        wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, "body > div.gx-mask.gx-unmask")))
                        driver.find_element(By.CSS_SELECTOR, "#Tab_TAB1Containerpanel4").click()
                        combo = driver.find_element(By.CSS_SELECTOR, "#W0038vACTFUENTADMINID")
                        select = Select(combo)
                        select.select_by_index(1) 
                        archivopropio = indice_folio+str(datos["Folio"])
                        try:
                            time.sleep(1)
                            patron_id = re.compile(r'span_W0038ACTDOCUMENTODESC_\d+')
                            time.sleep(1)
                            elemento1 = driver.find_elements(By.XPATH,"//*[contains(@id, 'span_W0038ACTDOCUMENTODESC_')]")
                            pdfs = len(elemento1)
                            # Obtener el texto de los elementos
                            if pdfs == 2 and ((archivopropio + " B.pdf" == elemento1[0].text) or (archivopropio + " B.pdf" == elemento1[1].text)) and ((archivopropio + " J.pdf" == elemento1[0].text) or (archivopropio + " J.pdf" == elemento1[1].text)):
                                print ("PDF subidos, ",datos["Folio"])
                            else:
                                borrar_archivos(archivopropio,pdfs)
                                time.sleep(1)
                                subir_documentos(carpeta_almacenamiento,archivopropio) 
                                cambiosQC.value = cambiosQC.value + '\nNombres de archivos cambiados'
                                 
                        except: 
                            time.sleep(1)
                            subir_documentos(carpeta_almacenamiento,archivopropio)
                            cambiosQC.value = cambiosQC.value + '\nSin documentos, subidos'
                        
                        lista_cedulas = []
                        lista_bool = []
                        lista_tipoid = []
                        lista_porcen = []
                        sumatoria = 0
                        firstced = True #no se usa
                        unaveznomas = True               
                        primero = False
                        sininteres = True
                        cedula_sinborrar = ""
                        tipo_idsinborrar = ""  
                            
                        time.sleep(.4)
                        driver.find_element(By.CSS_SELECTOR,'#Tab_TAB1Containerpanel7').click()
                        time.sleep(.4)
                        try:
                            botoncito = driver.find_element(By.CSS_SELECTOR, "#W0062GRIDPAGINATIONBARContainer_DVPaginationBar > div.PaginationBarCaption.dropdown > div > button")
                            botoncito.send_keys(Keys.PAGE_DOWN)
                            time.sleep(.6)
                            botoncito.click()
                            driver.find_element(By.XPATH, '//*[@id="W0062GRIDPAGINATIONBARContainer_DVPaginationBar"]/div[2]/div/ul/li[6]/a/span').click()
                            wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, '#W0062GridContainerRow_0001'))) 
                            time.sleep(.6)
                            driver.execute_script("window.scrollTo(0,0)")
                            
                        except:
                            print ("Sin interesados registrados, ",datos["Folio"])
                            sininteres = False
                            primero = True  

                        # Lista de claves correspondientes a los nombres y apellidos
                        nombres_apellidos_claves = ['Primer Nombre', 'Segundo Nombre', 'Primer Apellido', 'Segundo Apellido']
                        
                        if revisar_interesados == True:
                            interes_antes = ""
                            for i in range(cedulas_revisar):  
                                # Itera a través de las claves y modifica los valores in situ
                                if datosjuridicos["Género"] == 'N':
                                    priape = datosjuridicos["Primer Apellido"] 
                                    segape = datosjuridicos["Segundo Apellido"]
                                    
                                    datosjuridicos["Segundo Apellido"] = datosjuridicos["Segundo Nombre"]
                                    if datosjuridicos["Primer Nombre"] == '':
                                        datosjuridicos["Primer Apellido"] = segape
                                    else:
                                        datosjuridicos["Primer Apellido"] = datosjuridicos["Primer Nombre"]
                                        datosjuridicos["Segundo Nombre"] = segape
                                    
                                    datosjuridicos["Primer Nombre"] = priape  
                                        
                                for clave in nombres_apellidos_claves:
                                    if clave in ['Primer Nombre', 'Segundo Nombre'] and datosjuridicos[clave] != "":
                                        datosjuridicos[clave] = datosjuridicos[clave] + " "
                                    elif clave in ['Primer Apellido', 'Segundo Apellido'] and datosjuridicos[clave] != "" and datosjuridicos['Segundo Apellido'] != "" and clave != 'Segundo Apellido':
                                        datosjuridicos[clave] = datosjuridicos[clave] + " "
                                
                                #--------------------------------------------
                                buscar_malos,cedula_sinborrar,cadenanamemala = buscar_inter_malo(driver, datosjuridicos, sininteres, unaveznomas,i,cedulas_revisar)
                                #--------------------------------------------
                                #buscar_malos[2] = True
                                
                                
                                if modoqcl or modoqc and any(buscar_malos):
                                
                                    if buscar_malos[2]:
                                        unaveznomas = False
                                        firstced = False
                                        if pregunto_derechos == False:
                                            driver.find_element(By.CSS_SELECTOR, "#Tab_TAB1Containerpanel8").click()
                                            try:
                                                espere = WebDriverWait(driver, 2)
                                                espere.until(EC.visibility_of_element_located((By.CSS_SELECTOR, 'img[id^="W0070vDELETE_"]')))
                                                borrar_dere = driver.find_elements(By.CSS_SELECTOR,'img[id^="W0070vDELETE_"]')       #BORRAR DERECHOS
                                                # Itera a través de los botones y haz clic en cada uno de ellos
                                                for boton in borrar_dere:
                                                    wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, "body > div.gx-mask.gx-unmask")))
                                                    boton.click()
                                            except:
                                                print ("Sin derechos que eliminar, ", datos["Folio"])
                                            pregunto_derechos = True
                                            
                                        if sininteres:
                                            panel = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, '#Tab_TAB1Containerpanel6')))
                                            panel.click()
                                            
                                            time.sleep(1.5)
                                            wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, 'img[id^="W0054vDELETE2_"]')))
                                            time.sleep(.5)
                                            driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
                                            time.sleep(.5)
                                            driver.find_element(By.CSS_SELECTOR, "#W0054GRIDPAGINATIONBARContainer_DVPaginationBar > div.PaginationBarCaption.dropdown > div > button").click()
                                            driver.find_element(By.CSS_SELECTOR, "#W0054GRIDPAGINATIONBARContainer_DVPaginationBar > div.PaginationBarCaption.dropdown > div > ul > li:nth-child(6) > a > span").click()
                                            
                                            time.sleep(.5)
                                            driver.execute_script("window.scrollTo(0,0)")
                                            
                                            borrar = driver.find_elements(By.CSS_SELECTOR,'img[id^="W0054vDELETE2_"]')         #BORRAR DOCUMENTOS
                                            # Itera a través de los botones y haz clic en cada uno de ellos
                                            for boton in borrar:
                                                wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, "body > div.gx-mask.gx-unmask")))
                                                wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, 'img[id^="W0054vDELETE2_"]')))
                                                time.sleep(1.3)
                                                lolo = driver.find_element(By.CSS_SELECTOR,'img[id^="W0054vDELETE2_"]')
                                                time.sleep(1)
                                                lolo.click()
                                            sininteres = False
                                    
                                    if any(buscar_malos):
                                        if buscar_malos[1] and cedula_sinborrar != '0':
                                            ced_encontrada = cedula_sinborrar
                                        else:
                                            ced_encontrada = ''
                                        # if 'CARLOS' in datosjuridicos["Primer Nombre"]:
                                        #     print ("hola")
                                        
                                        r_cedula,r_bool = crear_interes(datosjuridicos["Primer Nombre"],
                                                            datosjuridicos["Segundo Nombre"],
                                                            datosjuridicos["Primer Apellido"],
                                                            datosjuridicos["Segundo Apellido"],
                                                            datosjuridicos["Género"],
                                                            i+1,
                                                            datosjuridicos["Cédulas"],
                                                            ced_encontrada,
                                                            )
                                        
                                        lista_cedulas.append(r_cedula)
                                        lista_bool.append(r_bool)
                                        lista_tipoid.append(datosjuridicos["Género"])
                                        lista_porcen.append(datosjuridicos["Porcentajes"])
                                
                                if any(buscar_malos):
                                    interes_antes = interes_antes + f"\n{cadenanamemala}"    
                                else:
                                    lista_cedulas.append(cedula_sinborrar)
                                    lista_bool.append(False)
                                    lista_tipoid.append(datosjuridicos["Género"])
                                    lista_porcen.append(datosjuridicos["Porcentajes"])
                                    
                                # Extrae los datos de la fila y almacénalos en el diccionario
                                fila_juridico+= 1
                                filaj = hojajuridico[fila_juridico]#tonto
                                for encabezado, columna in datos_titulosjuri.items():
                                    if columna:
                                        if filaj[columna-1].value == None:
                                            datosjuridicos[encabezado] = ""
                                        elif isinstance(filaj[columna-1].value, str):
                                            datosjuridicos[encabezado] = filaj[columna-1].value.strip()
                                        else:
                                            datosjuridicos[encabezado] = filaj[columna-1].value
                            
                            if modoqcl or not unaveznomas:    
                                firstced = True
                                i = 0
                                if unaveznomas == False:
                                    cambiosQC.value = cambiosQC.value + f'\nInteresados modificados y agregados, por estos: \n {interes_antes}'
                                    for ced, ipid, porc in zip(lista_cedulas,lista_tipoid,lista_porcen):
                                        time.sleep(.5)
                                        papu = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, '#Tab_TAB1Containerpanel6')))
                                        papu.click()
                                        time.sleep(.5)
                                        #wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "#span_W0014vACTPREDIOORIGENID")))
                                        
                                        
                                        wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, '#W0054CREARGRUPO')))
                                        
                                        # # Busca el tr por su ID específico
                                        # tb_element = driver.find_element(By.ID,'W0054GridgruposContainerTbl')

                                        # # Encuentra todos los td dentro del tr
                                        # tr_elements = tb_element.find_element(By.CSS_SELECTOR,'tr')
                                        
                                        while True:
                                            rows = driver.find_elements(By.CSS_SELECTOR, '[id^="W0054GridgruposContainerRow_"]')
                                            if len(rows) <= 1:
                                                break
                                            time.sleep(1)
                                            delete_button = driver.find_element(By.CSS_SELECTOR,'#W0054vDELETE1_0001')
                                            delete_button.click()
                                            time.sleep(.5)
                                            wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, "body > div.gx-mask")))
                                            wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, "body > div.gx-mask.gx-unmask")))
                                            # tr_elements = tb_element.find_element(By.CSS_SELECTOR,'td')  # Actualiza la lista de td después de eliminar
                                        
                                        while True:             #para asegurarse de que hay un grupo creado
                                            try:
                                                driver.find_element(By.CSS_SELECTOR, "#W0054GridgruposContainerRow_0001 > td:nth-child(4) > p")
                                                time.sleep(.5)
                                                grupo = driver.find_element(By.CSS_SELECTOR, "#W0054vSELECCIONAR_0001")      # escoger grupo
                                                time.sleep(.5)
                                                grupo.click()   
                                                time.sleep(.5)
                                                break
                                            except:
                                                driver.find_element(By.CSS_SELECTOR, "#W0054CREARGRUPO").click()
                                                time.sleep(1.5)
                                        
                                        if ipid != "N":
                                            driver.find_element(By.CSS_SELECTOR, "#W0054vTIPOIDENTIFICACION_ID").send_keys("Cedula Ciudadanía")
                                        else:
                                            driver.find_element(By.CSS_SELECTOR, "#W0054vTIPOIDENTIFICACION_ID").send_keys("Numero de Identificación Tributaria")
                                        time.sleep(1)
                                        if firstced:
                                            checkpro = driver.find_element(By.CSS_SELECTOR, "#W0054vACTGRUPINTEREPROPRI")      # escoger grupo
                                            checkpro.click()
                                            firstced = False
                                            time.sleep(1)
                                        
                                        driver.find_element(By.CSS_SELECTOR, "#W0054vACTMIENINTENUMDOC").send_keys(ced)
                                        
                                        if porc != None and porc != "":
                                            if "=" in str(porc):
                                                porc = porc.replace('=','')
                                                porc = eval(porc)
                                            valor_porcentaje = porc
                                        else:                                        
                                            valor_porcentaje =100/cedulas_revisar
                                            
                                        # valor_porcentajestr = str(valor_porcentaje)
                                        # valor_porcentajestr = valor_porcentajestr.replace(".",",")
                                        
                                        participacion = driver.find_element(By.CSS_SELECTOR, "#W0054vACTGRUINTEREPARTIC")
                                        participacion.clear()
                                        participacion.send_keys(valor_porcentaje)
                                        
                                        driver.find_elements(By.CSS_SELECTOR, "ody > div.ui-pnotify.stack-topright > div")
                                        
                                        wait.until(lambda driver: len(driver.find_elements(By.CSS_SELECTOR, ".ui-pnotify.stack-topright[style*='display: block']")) <= 2)
                                        driver.find_element(By.CSS_SELECTOR, "#W0054ASOCIARINTERESADO").click()
                                        wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, "body > div.gx-mask.gx-unmask"))) 
                                        wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "body > div.ui-pnotify.stack-topright > div > div.ui-pnotify-text"))) 
                                        time.sleep(1) 
                                        count = 0
                                        if i == cedulas_revisar - 1:
                                            
                                            while True:
                                                try:
                                                    
                                                    notificacion = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "body > div.ui-pnotify.stack-topright > div > div.ui-pnotify-text")))  
                                                    notitext = notificacion.text
                                                    texto = "No es posible agregar el registro ya que la participación total entre los interesados sumaría más de 100..."
                                                    texto2 = "Los datos han sido agregados..."
                                                    if notitext == texto:
                                                        wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, "body > div.ui-pnotify.stack-topright > div > div.ui-pnotify-text")))
                                                        raise Exception('ERROR: No se pudieron agregar todos los interesados')
                                                    else:
                                                        break
                                                except:
                                                    count +=1
                                                    if count >= 6:
                                                        print ("ALERTAR, ULTIMO INTERESADO NO PUDO SER INCLUIDO, ",datos["Folio"])
                                                        wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, "body > div.ui-pnotify.stack-topright > div > div.ui-pnotify-text")))
                                                        break
                                                    valor_porcentaje = float(valor_porcentaje) - .1
                                                    driver.find_element(By.CSS_SELECTOR, "#W0054vACTGRUINTEREPARTIC").send_keys(str(valor_porcentaje).replace(".",","))
                                                    driver.find_element(By.CSS_SELECTOR, "#W0054ASOCIARINTERESADO").click()
                                                    wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, "body > div.gx-mask.gx-unmask"))) 
                                        sumatoria = sumatoria + float(valor_porcentaje)
                                    
                                
                                
                                if not (95 < sumatoria < 100):
                                    print ("ALERTA: La suma de los interesados no está entre 95 y 100 % ---> ", sumatoria)
                                
                                falta_cedula = False
                                if unaveznomas == True and buscar_malos[1]:
                                    falta_cedula = True
                                
                                if any(lista_bool) and unaveznomas == False or falta_cedula:
                                    set_zero_intereados(lista_tipoid,lista_cedulas,lista_bool)
                                
                                
                                time.sleep(.4)
                                driver.find_element(By.CSS_SELECTOR,'#Tab_TAB1Containerpanel7').click()
                                time.sleep(.4)
                                
                                
                                try:
                                    botoncito = driver.find_element(By.CSS_SELECTOR, "#W0062GRIDPAGINATIONBARContainer_DVPaginationBar > div.PaginationBarCaption.dropdown > div > button")
                                    botoncito.send_keys(Keys.PAGE_DOWN)
                                    time.sleep(.6)
                                    botoncito.click()
                                    driver.find_element(By.XPATH, '//*[@id="W0062GRIDPAGINATIONBARContainer_DVPaginationBar"]/div[2]/div/ul/li[6]/a/span').click()
                                    wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, '#W0062GridContainerRow_0001'))) 
                                    time.sleep(.6)
                                    driver.execute_script("window.scrollTo(0,0)")
                                except:
                                    print ("Sin interesados registrados, ",datos["Folio"])
                                    sininteres = False
                                    primero = True  

                                
                                for i in range(cedulas_revisar):  
                                    #--------------------------------------------
                                    if i != 0:
                                        fila_juridico_inicial+= 1
                                    filaj = hojajuridico[fila_juridico_inicial]#tonto
                                    for encabezado, columna in datos_titulosjuri.items():
                                        if columna:
                                            if filaj[columna-1].value == None:
                                                datosjuridicos[encabezado] = ""
                                            elif isinstance(filaj[columna-1].value, str):
                                                datosjuridicos[encabezado] = filaj[columna-1].value.strip()
                                            else:
                                                datosjuridicos[encabezado] = filaj[columna-1].value
                                                
                                    if datosjuridicos["Género"] == 'N':
                                        priape = datosjuridicos["Primer Apellido"] 
                                        segape = datosjuridicos["Segundo Apellido"]
                                        datosjuridicos["Segundo Apellido"] = datosjuridicos["Segundo Nombre"]
                                        if datosjuridicos["Primer Nombre"] == '':
                                            datosjuridicos["Primer Apellido"] = segape
                                        else:
                                            datosjuridicos["Primer Apellido"] = datosjuridicos["Primer Nombre"]
                                            datosjuridicos["Segundo Nombre"] = segape
                                        
                                        datosjuridicos["Primer Nombre"] = priape                                
                                                
                                    for clave in nombres_apellidos_claves:
                                        if clave in ['Primer Nombre', 'Segundo Nombre'] and datosjuridicos[clave] != "":
                                            datosjuridicos[clave] = datosjuridicos[clave] + " "
                                        elif clave in ['Primer Apellido', 'Segundo Apellido'] and datosjuridicos[clave] != "" and datosjuridicos['Segundo Apellido'] != "" and clave != 'Segundo Apellido':
                                            datosjuridicos[clave] = datosjuridicos[clave] + " "
                                            
                                    buscar_malos,poop, papa= buscar_inter_malo(driver, datosjuridicos, True, True,i,cedulas_revisar)
                                    if any(buscar_malos):
                                        raise Exception('ERROR: Última comprobación de interesados, fallida')
                        #if modoqc:
                        if llenar_derechos:
                            derechos(datosjuridicos["Texto"])
                        wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, "body > div.gx-mask.gx-unmask")))
                        driver.find_element(By.CSS_SELECTOR, "#Tab_TAB1Containerpanel1").click()
                        time.sleep(0.5)
                        driver.find_element(By.CSS_SELECTOR, "#W0014CANCEL").click()
                    flag = 0
                    errorzaso.value = None
                    
                    archivo_excel.save(carpeta_almacenamiento+"Libro1.xlsx")
            break
    
        except Exception as e:
            mensaje_error = traceback.format_exc()
            a += 1
            print(f"Fila: {a}, IteracionSe produjo un error: {e}")
            flag = 0
            error.value = "NO SE LLENÓ FOLIO, REVISAR"
            
            errorzaso.value = f"Se produjo un error: {e}"
            
    
    #print ("FOLIO COMPLETADO!!!!!! ----> ", datos["Folio"])
    fila_a_extraer+=1
        

# Restaurar la configuración de suspensión
ctypes.windll.kernel32.SetThreadExecutionState(0x80000000)

reproducir_audio(ruta_audio)
time.sleep(30)

if HibernarPC == True:
    time.sleep(5)
    # Suspender el PC (este comando es específico para Windows)
    os.system("rundll32.exe powrprof.dll,SetSuspendState 0,1,0")
