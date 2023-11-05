from selenium import webdriver
from selenium.webdriver.common.by import By  # Importa el módulo By
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
import openpyxl
import time
import sys
import io
from bs4 import BeautifulSoup
import re

llenarelresto = 1               #LLENAR SOLO INFO DEL PREDIO O TAMBIEN EL RESTO DEL PROCESO????????     0 para solo predio, 1 para todo
nivel_de_zoom = 0.8
# Configura las opciones del navegador para ajustar el nivel de zoom
options = webdriver.ChromeOptions()
options.add_argument(f"--force-device-scale-factor={nivel_de_zoom}")

# Crea una instancia de un navegador web (por ejemplo, Edge)
driver = webdriver.Chrome(options=options)
# Maximiza la ventana del navegador
driver.maximize_window()
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# Obtener el contenido HTML de la página
html = driver.page_source
# Crear un objeto BeautifulSoup para analizar el HTML
soup = BeautifulSoup(html, "html.parser")

indice_folio = "148-"
# Abre el archivo Excel
carpeta_almacenamiento= 'C:/Users/PORTATIL LENOVO/Downloads/Pruebas_autom/31-10-2023/'
nombre_excel = 'Libro1.xlsx'
archivo_excel = openpyxl.load_workbook(carpeta_almacenamiento+nombre_excel)

# Selecciona la hoja en la que deseas trabajar
hoja = archivo_excel['informacion_propiedades']
hojajuridico = archivo_excel['nombres_cedulas']

def quitar_acentos(texto):
    acentos = {
        'á': 'a', 'é': 'e', 'í': 'i', 'ó': 'o', 'ú': 'u',
        'Á': 'A', 'É': 'E', 'Í': 'I', 'Ó': 'O', 'Ú': 'U'
    }
    
    # Reemplaza caracteres acentuados
    for acento, sin_acento in acentos.items():
        texto = texto.replace(acento, sin_acento)
    
    return texto

def crear_interes(pn,sn,pa,sa,genre,conteo_ced,cedula=""):
#def crear_interes(pn,pa,genre):
    # Abre el archivo de Excel R1
    archivo_excel = openpyxl.load_workbook('C:/Users/PORTATIL LENOVO/Downloads/R1.xlsx')
    # Selecciona la hoja en la que deseas buscar
    hoja = archivo_excel['Hoja1']
    archivo_excel.close()
    
    wait = WebDriverWait(driver, 10)
    panel = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "#Tab_TAB1Containerpanel6")))
    panel.click()
    time.sleep(1.5)
    paneol = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "#W0054CREARINTERESADO")))
    time.sleep(1.5)
    wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, "body > div.gx-mask.gx-unmask")))
    paneol.click()
    time.sleep(1.5)


    persona = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "#vSOLICITANTE_ID")))
    if genre != "N":                #N: NIT, H: hombre, M: mujer
        posicion_id = 0                                     #posicion del ciclo en el while
        tipo_id_selecc = False                                 #En qué tipo de id fue encontrado el interesado
        fix_ced = False
        fix_gen = False
        same_nombres = 0                    #cuantos nombres iguales se encontraron
        tipo_id_selecc_str = ""
        cedula_selecc = ""
        editar_selecc = ""
        while True:
            persona.send_keys("PERSONA NATURAL")
            
            vernombre = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "#vACTMIENINTEPNOMBRE")))
            vernombre.clear()
            vernombre.send_keys(pn)
            pepo = driver.find_element(By.CSS_SELECTOR, "#vACTMIENINTEPAPELLIDO")
            pepo.clear()
            pepo.send_keys(pa)
            
            tipo_id = ["Cedula Ciudadanía","Tarjeta Identidad","Secuencial"]
            
            while True:
                
                try: 
                    driver.find_element(By.CSS_SELECTOR, "#vTIPOIDENTIFICACION_ID").send_keys(tipo_id[posicion_id])
                    
                    time.sleep(2)
                    if posicion_id == 2:
                        print ("hola")
                    if posicion_id == 0:
                        driver.find_element(By.CSS_SELECTOR, "#GRIDPAGINATIONBARContainer_DVPaginationBar > div.PaginationBarCaption.dropdown > div > button").click()
                        driver.find_element(By.XPATH, "//ul/li/a/span[contains(text(), 'All rows')]").click()
                        
                    time.sleep(1)
                    wait = WebDriverWait(driver, 10)
                    wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, "body > div.gx-mask.gx-unmask")))
                    
                    # Busca todos los elementos <tr> en la página
                    elementos_tr = driver.find_elements(By.TAG_NAME, "tr")
                    # print (elementos_tr)
                    #Itera a través de los elementos <tr>
                    cuantos = -1
                    same_nombres_ciclo = 0
                    
                    if "PATERNINA" in pa:
                        print ("hola")
                    
                    for elemento in elementos_tr:
                        # Busca el elemento td en la tercera posición (índice 2)
                        nombre_elemento = elemento.find_elements(By.CSS_SELECTOR, "td:nth-child(3) span")
                        if nombre_elemento:
                            nombre = quitar_acentos(nombre_elemento[0].text)
                            cuantos += 1
                            print(nombre)   #Saber qué nombres se enlistan
                            if nombre == (pn+" "+sn+" "+pa+" "+sa or pn+" "+pa+" "+sa):
                                tipo_id_selecc_str = tipo_id[posicion_id]                                               
                                same_nombres +=1
                                same_nombres_ciclo += 1
                                #REVISAR SI SE DEBE PONER EL same_nombres > 1: ACA
                                cedula_selecc = elemento.find_elements(By.CSS_SELECTOR, "td:nth-child(5) span")
                                cedula_selecc = cedula_selecc[0].text
                                genero_selecc = elemento.find_elements(By.CSS_SELECTOR, "td:nth-child(8) span")
                                genero_selecc = genero_selecc[0].text
                                editar_selecc = elemento.get_attribute("data-gxrow")
                                # Itera a través de las filas de la columna H
                                for fila in hoja.iter_rows(min_col=8, max_col=11, values_only=True):                     #COMPARAR CEDULA CON ALGUNA EN R1
                                    # Verifica si la cadena de texto que estás buscando se encuentra en la celda
                                    sin_acento = quitar_acentos(fila[0])
                                    if (pn+" "+sn+" "+pa+" "+sa or pn+" "+pa+" "+sa or " "+pn+" "+sn+" "+pa+" "+sa or " "+pn+" "+pa+" "+sa) in sin_acento:
                                        if cedula_selecc != (fila[3] and '0') and '-' not in cedula_selecc:
                                            fix_ced = True 
                                            cedula = fila[3]
                                    else:
                                        fix_ced = True
                                                            
                    if same_nombres_ciclo == 0:
                        raise Exception()
                    elif same_nombres > 1:
                        print ("MISMO NOMBRE ENCONTRADO DOS VECES, REVISAR: ", nombre_elemento)
                        break
                    else:
                        break
                except:
                    popi = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "#vSOLICITANTE_ID")))
                    # Crea un objeto Select para interactuar con el ComboBox
                    select = Select(popi)
                    # Obtén el valor actual seleccionado en el ComboBox
                    valor_actual = select.first_selected_option.text
                    if valor_actual == "Seleccione...":
                        driver.refresh()
                        break
                    
                    posicion_id += 1
                    if posicion_id > 2:
                        break
                    tipo_id_selecc = True
            
            posicion_id += 1
            if posicion_id > 2:
                break
                    
                            #print(nombre)   #Saber qué nombres se enlistan
                    #print (cuantos)         #Saber cuantos nombres hay listados
            
        cedula_cero = False
        if cedula == "" and fix_ced:                #para que si no hay cedula, se establezca una cedula 9 98 987 etc
                cedula_cero = True
                cedula = "9" * 2 + "9" * (conteo_ced - 1)
                    
        if genre == "H":
            genre = "MASCULINO"
        elif genre == "M":
            genre = "FEMENINO"          
                  
        if same_nombres != 0:
            wait = WebDriverWait(driver, 10)
            time.sleep(.4)
            driver.find_element(By.CSS_SELECTOR, "#vTIPOIDENTIFICACION_ID").send_keys(tipo_id_selecc_str)
            wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "#vUPDATE_"+editar_selecc))) # Ajusta el selector según tu página.
            wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, "body > div.gx-mask.gx-unmask")))            
            driver.find_element(By.CSS_SELECTOR, "#vUPDATE_"+editar_selecc).click()
            driver.find_element(By.CSS_SELECTOR, "#vUPDATE_"+editar_selecc).click()
            time.sleep(1.5)
            
            #wait.until(EC.text_to_be_present_in_element_value((By.CSS_SELECTOR, "#vACTMIENINTENUMDOC"), cedula_selecc))  
            
            if fix_ced:
                driver.find_element(By.CSS_SELECTOR, "#vACTMIENINTENUMDOC").clear()
                driver.find_element(By.CSS_SELECTOR, "#vACTMIENINTENUMDOC").send_keys(cedula)
                wait.until(EC.text_to_be_present_in_element_value((By.CSS_SELECTOR, "#vACTMIENINTENUMDOC"), str(cedula)))  
            driver.find_element(By.CSS_SELECTOR, "#vACTGENEROID").send_keys(genre)
            driver.find_element(By.CSS_SELECTOR, "#vACTGRUPOETNICOID").send_keys("Ninguno")
            time.sleep(2)
            driver.find_element(By.CSS_SELECTOR, "#vTIPOIDENTIFICACION_ID").send_keys(tipo_id[0])
            # if fix_gen:
            #     driver.find_element(By.CSS_SELECTOR, "#vACTGENEROID").send_keys(genre)
            # if tipo_id_selecc:
            #     driver.find_element(By.CSS_SELECTOR, "#vTIPOIDENTIFICACION_ID").send_keys(tipo_id[0])
        else:
            driver.find_element(By.CSS_SELECTOR, "#vACTMIENINTENUMDOC").send_keys(cedula)
            time.sleep(.8)
            driver.find_element(By.CSS_SELECTOR, "#vACTGENEROID").send_keys(genre)
            time.sleep(.4)
            driver.find_element(By.CSS_SELECTOR, "#vTIPOIDENTIFICACION_ID").send_keys(tipo_id[0])
            driver.find_element(By.CSS_SELECTOR, "#vACTGRUPOETNICOID").send_keys("Ninguno")
            driver.find_element(By.CSS_SELECTOR, "#vACTMIENINTEPNOMBRE").clear()
            driver.find_element(By.CSS_SELECTOR, "#vACTMIENINTEPNOMBRE").send_keys(pn)
            driver.find_element(By.CSS_SELECTOR, "#vACTMIENINTEPAPELLIDO").clear()
            driver.find_element(By.CSS_SELECTOR, "#vACTMIENINTEPAPELLIDO").send_keys(pa)
            driver.find_element(By.CSS_SELECTOR, "#vACTMIENINTESNOMBRE").send_keys(sn)
            driver.find_element(By.CSS_SELECTOR, "#vACTMIENINTESAPELLIDO").send_keys(sa)
        
    else:
        cedula_cero = False
        persona.send_keys("PERSONA JURIDICA")
        driver.find_element(By.CSS_SELECTOR, "#vTIPOIDENTIFICACION_ID").send_keys("Numero de Identificación Tributaria")
        driver.find_element(By.CSS_SELECTOR, "#vACTMIENINTENUMDOC").send_keys(cedula)
        time.sleep(2)
        elemento = driver.find_element(By.CSS_SELECTOR, "#GRIDPAGINATIONBARContainer_DVPaginationBar > div")
        if "PaginationBarEmptyGrid" not in elemento.text:
            if cedula == "":                #para que si no hay cedula, se establezca una cedula 9 98 987 etc
                cedula_cero = True
                cedula = "9" * 2 + "9" * (conteo_ced - 1)

            driver.find_element(By.CSS_SELECTOR, "#vACTMIENINTERAZSOC").send_keys(pa+" "+sa+" "+pn+" "+sn)
            driver.find_element(By.CSS_SELECTOR, "#vACTMIENINTEDIGIVERI").send_keys("0")
            driver.find_element(By.CSS_SELECTOR, "#vACTMIENINTENUMDOC").send_keys(cedula)
            
        time.sleep(2)
        driver.find_element(By.CSS_SELECTOR, "#GUARDAR").click()
        time.sleep(2)
        wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, "body > div.gx-mask.gx-unmask")))
        # else:
        #     for fila in hoja.iter_rows(min_col=8, max_col=11, values_only=True):                     #COMPARAR CEDULA CON ALGUNA EN R1
        #         # Verifica si la cadena de texto que estás buscando se encuentra en la celda
        #         sin_acento = quitar_acentos(fila[0])
        #         if (pn+" "+sn+" "+pa+" "+sa or " "+pn+" "+sn+" "+pa+" "+sa) in sin_acento:
        #             if cedula_selecc != fila[3]:
        #                 fix_ced = True 
        #                 cedula = fila[3]     
    time.sleep(2)
    driver.find_element(By.CSS_SELECTOR, "#GUARDAR").click()
    time.sleep(2)
    wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, "body > div.gx-mask.gx-unmask")))
    papa = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "#CANCEL")))
    papa.click()
    
    return cedula,cedula_cero
        
def llenar_predio(hojainfo,mat_matriz,direccion,areaTerr):
    
    matricula = driver.find_element(By.CSS_SELECTOR, "#Tab_TAB1Containerpanel1")
    matricula.click()
    time.sleep(1)
    
    informacion_combobox = {
        '#W0014vACTPREDIOMATRICULAMATR': mat_matriz,
        '#W0014vACTPREDIODIREPRIP': direccion,
        '#W0014vACTPREDIOARETERR': areaTerr,
    }     

    for combobox_id, opcion in informacion_combobox.items():
        
        element = driver.find_element(By.CSS_SELECTOR, combobox_id)
        if combobox_id == '#W0014vACTPREDIODIREPRIP':
            valor = element.get_attribute("value")
            if valor.strip() and opcion is None:
                print ("Direccion no cambiada")
            else:  
                element.clear()
                time.sleep(.5)
                element.send_keys(opcion)
        else:
            element.clear()
            time.sleep(.5)
            element.send_keys(opcion)

    checkbox_element = driver.find_element(By.CSS_SELECTOR, "#W0014vACT_PREDIOAREAREGISTRAL")      # BOTON DE QUE NO HAY AREA CATASTRAL
    if not checkbox_element.is_selected():
        checkbox_element.click()
    # Encuentra el elemento por su ID
    elemento = driver.find_element(By.CSS_SELECTOR, "#W0014vACTPREDIOARECONST")

    # Añade un nuevo dato a la celda a la derecha de "Dirección Corregida"
    nueva_celda = hojainfo.cell(row=fila_a_extraer, column=columna_max + 1)
    nueva_celda.value = elemento.get_attribute("value")  # Reemplaza con el valor que desees
    
    driver.find_element(By.CSS_SELECTOR, "#W0014GUARDAR").click()
    
    time.sleep(1)
    wait = WebDriverWait(driver, 10)
    wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, "body > div.gx-mask.gx-unmask")))



def crear_fuente(fuente,date_doc,ente,date_reg,n_fuente="SN"):

    matricula = driver.find_element(By.CSS_SELECTOR, "#Tab_TAB1Containerpanel3")
    matricula.click()
    time.sleep(1)
    
    date_doc = date_doc.strftime("%d/%m/%Y")
    date_reg = date_reg.strftime("%d/%m/%Y")
    # Define un diccionario con los selectores CSS y sus valores
    campos = {
        "#W0030vACTFUENTEADMINTIPOID": fuente,
        "#W0030vACTPRINCIPALDOCTIPOID": "Documento",
        "#W0030vACTNUMFUENTE": n_fuente,
        "#W0030vACTENTEEMISOR": ente
    }

    # Llena los campos utilizando un bucle
    for selector, valor in campos.items():
        campo = driver.find_element(By.CSS_SELECTOR, selector)
        campo.send_keys(valor)

    # driver.find_element(By.CSS_SELECTOR, "#W0030vACTFUEADMFECHDOC").send_keys(date_doc.strftime("%d/%m/%Y"))
    # driver.find_element(By.CSS_SELECTOR, "#W0030vACTFUEADMFECHREG").send_keys(date_reg.strftime("%d/%m/%Y"))
    # Completa las fechas utilizando driver.execute_script
    driver.execute_script("arguments[0].value = arguments[1];", driver.find_element(By.CSS_SELECTOR, "#W0030vACTFUEADMFECHDOC"), date_doc)
    driver.execute_script("arguments[0].value = arguments[1];", driver.find_element(By.CSS_SELECTOR, "#W0030vACTFUEADMFECHREG"), date_reg)

    matricula = driver.find_element(By.CSS_SELECTOR, "#W0030ENTER")
    matricula.send_keys(Keys.PAGE_DOWN)
    matricula.click()
    
    time.sleep(1)
    wait = WebDriverWait(driver, 15)
    wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, "body > div.ui-pnotify.stack-topright")))
    
def subir_documentos(ruta,cadena_folio):
    driver.find_element(By.CSS_SELECTOR, "#Tab_TAB1Containerpanel4").click()
    combo = driver.find_element(By.CSS_SELECTOR, "#W0038vACTFUENTADMINID")
    select = Select(combo)
    select.select_by_index(1)
    
    file_input = driver.find_element(By.CSS_SELECTOR,"#W0038FILEUPLOAD1Container > div > div.row.fileupload-buttonbar > div.col-lg-7 > span.btn.btn-success.fileinput-button > input[type=file]")  # Reemplaza "fileInput" con el ID de tu elemento

    # Usa send_keys() para proporcionar la ruta completa del archivo a cargar
    time.sleep(1.5)
    file_input.send_keys(ruta+cadena_folio+" J.pdf")
    #time.sleep(1)
    file_input = driver.find_element(By.CSS_SELECTOR,"#W0038FILEUPLOAD1Container > div > div.row.fileupload-buttonbar > div.col-lg-7 > span.btn.btn-success.fileinput-button > input[type=file]")  # Reemplaza "fileInput" con el ID de tu elemento
    time.sleep(1.5)
    file_input.send_keys(ruta+cadena_folio+" B.pdf")
    driver.find_element(By.CSS_SELECTOR, "#W0038FILEUPLOAD1Container > div > div.row.fileupload-buttonbar > div.col-lg-7 > button.btn.btn-primary.start").click()
    wait = WebDriverWait(driver, 10)
    time.sleep(3)
    wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, "#W0038FILEUPLOAD1Container > div > table > tbody > tr:nth-child(1) > td:nth-child(4) > button.btn.btn-primary.start")))
    
    driver.find_element(By.CSS_SELECTOR, "#W0038CONFIRMAR").click()
    time.sleep(1.5)
    wait = WebDriverWait(driver, 10)
    wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, "body > div.ui-pnotify.stack-topright")))
 
def derechos():
    wait = WebDriverWait(driver, 10)
    pato = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "#Tab_TAB1Containerpanel8")))
    pato.click()

    selectores = ["#W0070vACTFUENTADMINID", "#W0070vACTGRUPINTEREID", "#W0070vACTDERECHOTIPOID"]

    for selector in selectores:
        combo = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, selector)))
        combo = driver.find_element(By.CSS_SELECTOR, selector)
        select = Select(combo)
        wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, f"{selector} option:nth-child({1 + 1})")))
        time.sleep(3)
        select.select_by_index(1)

    driver.find_element(By.CSS_SELECTOR, "#W0070GUARDAR").click()
    wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, "body > div.gx-mask.gx-unmask")))
    
    
    
 
def set_zero_intereados(list_identificacion, lis_cedulas,list_booleanos):
    driver.find_element(By.CSS_SELECTOR, "#Tab_TAB1Containerpanel6").click()
    driver.find_element(By.CSS_SELECTOR, "#W0054CREARINTERESADO").click()
    
    wait = WebDriverWait(driver, 10)

    
    for identificadores, ceduloides, booleanos in zip(list_identificacion, lis_cedulas,list_booleanos):
        if booleanos == True:
            while True:
                try:
                    wait.until(lambda driver: len(driver.find_elements(By.CSS_SELECTOR, ".ui-pnotify.stack-topright[style*='display: block']")) <= 1)
                    driver.find_element(By.CSS_SELECTOR, "#LIMPIARCAMPOS").click()
                    try:
                        wait = WebDriverWait(driver, 10)
                        wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, "body > div.gx-mask.gx-unmask")))
                    except:
                        driver.refresh()    
                    
                    if identificadores != "N":                #N: NIT, H: hombre, M: mujer
                        time.sleep(2)
                        wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "#vSOLICITANTE_ID"))).send_keys("PERSONA NATURAL")
                        wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "#vACTMIENINTEPNOMBRE")))
                        time.sleep(1)
                        driver.find_element(By.CSS_SELECTOR, "#vTIPOIDENTIFICACION_ID").send_keys("Cedula Ciudadanía")
                    else:
                        time.sleep(2)
                        wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "#vSOLICITANTE_ID"))).send_keys("PERSONA JURIDICA")
                        wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "#vACTMIENINTERAZSOC")))
                        time.sleep(2)
                        driver.find_element(By.CSS_SELECTOR, "#vTIPOIDENTIFICACION_ID").send_keys("Numero de Identificación Tributaria")
                    
                    driver.find_element(By.CSS_SELECTOR, "#vACTMIENINTENUMDOC").send_keys(ceduloides)
                    wait.until(EC.text_to_be_present_in_element_value((By.CSS_SELECTOR, "#vACTMIENINTENUMDOC"), ceduloides))  
                    
                    time.sleep(1.5)
                    paco = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "#GRIDPAGINATIONBARContainer_DVPaginationBar > div.PaginationBarCaption.dropdown > div > button")))
                    paco.click()
                    driver.find_element(By.XPATH, "//ul/li/a/span[contains(text(), 'All rows')]").click()
                    
                    wait = WebDriverWait(driver, 10)
                    wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, "body > div.gx-mask.gx-unmask")))
                    time.sleep(5)
                    # Busca todos los elementos <tr> en la página
                    elementos_tr = driver.find_elements(By.TAG_NAME, "tr")


                    for elemento in elementos_tr:
                        # Busca el elemento td en la tercera posición (índice 2)
                        nombre_elemento = elemento.find_elements(By.XPATH, ".//td[@data-colindex='4']//span")

                        if nombre_elemento:
                            nombre = nombre_elemento[0].text
                            if nombre == ceduloides:
                                editar_selecc = elemento.get_attribute("data-gxrow")
                                break
                
                    
                    pepo = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "#vUPDATE_"+editar_selecc))) # Ajusta el selector según tu página.
                    wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, "body > div.gx-mask.gx-unmask")))
                    time.sleep(1.5)
                    if identificadores == "N": 
                        pepo.send_keys(Keys.PAGE_DOWN)
                    time.sleep(1.5)
                    pepo.click()
                    time.sleep(1.5)
                    pepo.click()
                    pepo.send_keys(Keys.PAGE_UP)
                    time.sleep(1)
                    if identificadores != "N":                #N: NIT, H: hombre, M: mujer
                        if identificadores == "H":
                            identificadores = "MASCULINO"
                        elif identificadores == "M":
                            identificadores = "FEMENINO"   
                        time.sleep(1.5)
                        driver.find_element(By.CSS_SELECTOR, "#vACTGRUPOETNICOID").send_keys("Ninguno")

                                
                        driver.find_element(By.CSS_SELECTOR, "#vACTGENEROID").send_keys(identificadores)
                    driver.find_element(By.CSS_SELECTOR, "#vACTMIENINTENUMDOC").clear()
                    time.sleep(.5)
                    driver.find_element(By.CSS_SELECTOR, "#vACTMIENINTENUMDOC").send_keys("0")
                    wait.until(EC.text_to_be_present_in_element_value((By.CSS_SELECTOR, "#vACTMIENINTENUMDOC"), "0"))  
                    time.sleep(1)
                    driver.find_element(By.CSS_SELECTOR, "#GUARDAR").click()
                    time.sleep(1)
                    driver.find_element(By.CSS_SELECTOR, "#GUARDAR").click()
                    wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, "body > div.gx-mask.gx-unmask")))
                    time.sleep(2)               
                    break
                except:
                    driver.refresh()
                    
    driver.find_element(By.CSS_SELECTOR, "#CANCEL ").click()
    wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, "body > div.gx-mask.gx-unmask")))
    time.sleep(1.5)

                
def borrar_archivos(archivolio,cantidad):

    wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, "body > div.gx-mask.gx-unmask")))
    driver.find_element(By.CSS_SELECTOR, "#Tab_TAB1Containerpanel4").click()
    combo = driver.find_element(By.CSS_SELECTOR, "#W0038vACTFUENTADMINID")
    select = Select(combo)
    select.select_by_index(1) 

    
    try:
        wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, "body > div.gx-mask.gx-unmask")))
        wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, 'img[id^="W0038vBORRAR_000"]')))
        borrar = driver.find_element(By.CSS_SELECTOR,'img[id^="W0038vBORRAR_0001"]')             #BORRAR DOCUMENTOS
        borrar.click()
        time.sleep(1.2)
        
        if cantidad == 2:
            wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, "body > div.gx-mask.gx-unmask")))
            bobo= wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, 'img[id^="W0038vBORRAR_000"]')))
            bobo.click()
    except: 
        print ("Archivos eliminados")
            
def buscar_inter_malo(driver, datosjuridicos, sininteres, unaveznomas):
    buscar_malos = [False, False, True, False]
    cedsinborrar = ""
    if sininteres and unaveznomas:
        wait = WebDriverWait(driver, 10)
        wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, "body > div.gx-mask.gx-unmask")))
        time.sleep(5)
        elementos_tr = driver.find_elements(By.TAG_NAME, "tr")

        for elemento in elementos_tr:
            buscar_nombre = elemento.find_elements(By.CSS_SELECTOR, "td:nth-child(7) span")

            if buscar_nombre:
                buscar_nombre = quitar_acentos(buscar_nombre[0].text)
                print(buscar_nombre)
                
                if buscar_nombre == (datosjuridicos["Primer Nombre"] + " " + datosjuridicos["Segundo Nombre"] + " " + datosjuridicos["Primer Apellido"] + " " + datosjuridicos["Segundo Apellido"] or datosjuridicos["Primer Nombre"] + " " + datosjuridicos["Primer Apellido"] + " " + datosjuridicos["Segundo Apellido"]):
                    buscar_malos[2] = False
                    buscar_tipo = elemento.find_elements(By.CSS_SELECTOR, "td:nth-child(5) span")
                    buscar_tipo = buscar_tipo[0].text
                    buscar_ced = elemento.find_elements(By.CSS_SELECTOR, "td:nth-child(6) span")
                    buscar_ced = buscar_ced[0].text
                    buscar_gen = elemento.find_elements(By.CSS_SELECTOR, "td:nth-child(10) span")
                    buscar_gen = buscar_gen[0].text
                    cedsinborrar = buscar_ced

                    if datosjuridicos["Género"] in ("H", "M") and buscar_tipo != "Cedula Ciudadanía" or datosjuridicos["Género"] == "N" and buscar_tipo != "Numero de Identificación Tributaria":
                        buscar_malos[0] = True

                    patron = r'\b(?:9|98|987|9876|98765|987654|9876543|98765432)\b'
                    coincidencias = re.findall(patron, buscar_ced)

                    if any(substring in buscar_ced for substring in ("-", '98765432')) and datosjuridicos["Género"] in ("H", "M") or coincidencias:
                        buscar_malos[1] = True

                    if datosjuridicos["Género"] in ("H", "M") and buscar_gen == "SIN_DETERMINAR":
                        buscar_malos[3] = True
                    break
    return buscar_malos,cedsinborrar

# Crea un diccionario para mapear los nombres de encabezados a variables
datos = {
    'Folio': None,
    'Matrícula matriz': None,
    'Área de Terreno': None,
    'Dirección Corregida': None,
}

datosjuridicos = {
    'Folio': None,
    'Fecha registro': None,
    'Fecha documento': None,
    'Fuente adm.': None,
    'N. Fuente': None,
    'Ente Em.': None,
    'Cédulas': None,
    'Primer Nombre': None,
    'Segundo Nombre': None,
    'Primer Apellido': None,
    'Segundo Apellido': None,
    'Género': None
}
# Obtén los valores de los encabezados en la primera fila
for columna in hoja.iter_cols(min_row=1, max_row=1):
    for celda in columna:
        if celda.value in datos:
            datos[celda.value] = celda.column
            columna_max = celda.column

for columna in hojajuridico.iter_cols(min_row=1, max_row=1):
    for celda in columna:
        if celda.value in datosjuridicos:
            datosjuridicos[celda.value] = celda.column
            columna_max_juridico = celda.column

# Navega a la página web deseada
driver.get("https://www.realidad5.com/realmultipropositosahagun/servlet/com.realmultipropositogam.iniciosesion")


while True:
    try:
        # Localiza el campo de texto de usuario por su selector CSS
        campo_usuario = driver.find_element(By.CSS_SELECTOR, "#vUSERNAME")
        campo_usuario.send_keys("jgonzalez")

        # Localiza el campo de texto de contraseña por su selector CSS
        campo_contrasena = driver.find_element(By.CSS_SELECTOR, "#vUSERPASSWORD")
        #campo_contrasena.send_keys("u9OPY2tnZs*71")   #SAN LUIS
        campo_contrasena.send_keys("hg1PXuk@9sx34")

        entrar = driver.find_element(By.CSS_SELECTOR, "#BTNENTER")
        entrar.click()
        wait = WebDriverWait(driver, 5)
        mas = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "#TEXTBLOCKTITLE_MPAGE")))
        break
    except:
        driver.refresh()

fila_a_extraer = 2  # Reemplaza con el número de fila deseado REVISAR 14 Y 15,16
veces_repetir_folio = 5
datos_titulos = datos.copy()
datos_titulosjuri = datosjuridicos.copy()
flag = 0

while hoja[fila_a_extraer] != None:
    for a in range(veces_repetir_folio):
        try:
            print ("Llenando la fila # ",fila_a_extraer,"...")
            # Extrae los datos de la fila y almacénalos en el diccionario
            fila = hoja[fila_a_extraer]
            for encabezado, columna in datos_titulos.items():
                if columna:
                    if fila[columna-1].value == None:
                        datos[encabezado] = ""
                    else:
                        datos[encabezado] = fila[columna-1].value
            
            if flag != 1:    
                driver.get("https://www.realidad5.com/realmultipropositosahagun/servlet/com.realmultipropositogam.wwfichprediact")
                wait = WebDriverWait(driver, 5)
                holi = WebDriverWait(driver, 10)
                while True:
                    try:
                        mas = holi.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "#ACTIONNEW"))) # Ajusta el selector según tu página.
                        mas.click()
                        break
                    except:
                        driver.refresh()
            wait = WebDriverWait(driver, 3)   
            time.sleep(1)
            wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "#vACTPREDIOESTADO")))
            driver.find_element(By.CSS_SELECTOR, "#vACTPREDIOESTADO").send_keys("Activo")
            matricula = driver.find_element(By.CSS_SELECTOR, "#vACTPREDIOMATRICULA")
            matricula.clear()
            matricula.send_keys(datos['Folio'])
            flag = 0
            time.sleep(2)
                            
            while True:
                try:
                    
                    matricula = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "#vUPDATE_0001")))
                    matricula.click()
                    flag = 0
                    break
                except:
                    wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "#vACTPREDIOESTADO")))
                    driver.find_element(By.CSS_SELECTOR, "#vACTPREDIOESTADO").send_keys("Realizado")
                    if flag == 1:
                        nueva_celda = hoja.cell(row=fila_a_extraer, column=columna_max+3)
                        nueva_celda.value = "Unfounded"
                        break
                    flag = 1
            nueva_celda1 = hoja.cell(row=fila_a_extraer, column=columna_max + 1)
            
            if nueva_celda1.value is not None and nueva_celda1.value != "NO SE LLENÓ FOLIO, REVISAR":
                flag = 1

            if flag != 1:
                time.sleep(5)
                wait = WebDriverWait(driver, 15)
                combobox = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "#W0014vTIPOLOTE_ID"))) # Ajusta el selector según tu página.

                # Obtiene el texto del elemento seleccionado.
                elemento_seleccionado = Select(combobox).first_selected_option
                texto_elemento = elemento_seleccionado.text

                # Define la cadena de caracteres que deseas buscar.
                cadena_busqueda = "Unidad_Predial"

                # Comprueba si la cadena de búsqueda está presente en el texto.
                if cadena_busqueda not in texto_elemento:
                    llenar_predio(hoja,datos['Matrícula matriz'],datos['Dirección Corregida'],datos['Área de Terreno'])
                else:
                    nueva_celda = hoja.cell(row=fila_a_extraer, column=columna_max+ 2)
                    nueva_celda.value = "Es PH, no se modifica"
                
                if llenarelresto:
                    folio_selec = datos['Folio']
                    fila_juridico = 0
                    fila_cedulas = 1
                    juridico = False
                    for filaj in hojajuridico.iter_rows(min_col=2, max_col=14, values_only=True):                     #revisar en que posicion esta el folio y la cantidad de interesados
                        # Verifica si la cadena de texto que estás buscando se encuentra en la celda
                        if juridico == False:
                            if str(folio_selec) in str(filaj[0]):
                                juridico = True
                                fila_juridico = fila_cedulas
                        
                        elif juridico == True and filaj[0] != None or filaj[12] == None:
                            break
                        fila_cedulas += 1    
                    cedulas_revisar = fila_cedulas - fila_juridico
                    fila_juridico_inicial = fila_juridico
                    if juridico == False:
                        print ("NO HUBO COINCIDENCIA FOLIO BASICO CON EL JURIDICO: ",datos['Folio'])
                    else:
                        # Extrae los datos de la fila y almacénalos en el diccionario
                        filaj = hojajuridico[fila_juridico]
                        for encabezado, columna in datos_titulosjuri.items():
                            if columna:
                                if filaj[columna-1].value == None:
                                    datosjuridicos[encabezado] = ""
                                else:
                                    datosjuridicos[encabezado] = filaj[columna-1].value
                    if juridico != False:
                        wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, "body > div.gx-mask.gx-unmask")))
                        time.sleep(1.2)
                        matricula = driver.find_element(By.CSS_SELECTOR, "#Tab_TAB1Containerpanel3")
                        matricula.click()
                        time.sleep(1)
                        
                        pregunto_derechos = False
                        holi = False
                        holo = False
                        try:
                            elemento_fuente = driver.find_element(By.CSS_SELECTOR, "#span_W0030ACTNUMFUENTE_0001")
                            holi = True
                        except:
                            print ("No se encontró fuente administrativa, ", datosjuridicos["Folio"])
                            holo = True
                            
                        mod_fuente = False
                        if holi and str(datosjuridicos["N. Fuente"]) != elemento_fuente.text:       #MODIFICAR FUENTE: PRIMERO SE REVISA FOLIO SI FUENTE ES DIFERNTE
                            nueva_celda = hoja.cell(row=fila_a_extraer, column=columna_max+ 3)      #SI ES DIFERENTE PREGUNTAR AL USUARIO POR MEDIO DE LA CELDA MISMA
                            if nueva_celda.value == "1":                                            #QUE DONDE ESTÁ "UNFOUNDED", SI EL USUARIO QUIERE QUE SE MODIFIQUE
                                mod_fuente = True                                                   #DEBE PONER UN 1 EN ESA MISMA CELDA
                                nueva_celda.value = "Fuente adm. Mod."
                            else:
                                nueva_celda.value = elemento_fuente.text
                        #a = 1/0
                        if mod_fuente:
                            
                            ###############BORRAR DOCUMENTOS .PDF Y DERECHOS ADMINIS ######################
                            # borrar_archivos(datos["Folio"],2)

                            # driver.find_element(By.CSS_SELECTOR, "#Tab_TAB1Containerpanel8").click()
                            # pregunto_derechos = True
                            # try:
                            #     wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, 'img[id^="W0070vDELETE_"]')))
                            # except:
                            #     print ("No hubieron derechos o no se pudieron localizar, ", datos["Folio"])     
                                
                            # borrar_dere = driver.find_elements(By.CSS_SELECTOR,'img[id^="W0070vDELETE_"]')       #BORRAR DERECHOS
                            # # Itera a través de los botones y haz clic en cada uno de ellos
                            # for boton in borrar_dere:
                            #     wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, "body > div.gx-mask.gx-unmask")))
                            #     time.sleep(.4)
                            #     boton.click()

                            driver.find_element(By.CSS_SELECTOR, "#Tab_TAB1Containerpanel3").click()
                            wait = WebDriverWait(driver, 10)
                            borrar_fuente = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "#W0030vUPDATE_0001")))
                            borrar_fuente.click()
                                
                            crear_fuente(datosjuridicos["Fuente adm."],
                                        datosjuridicos["Fecha documento"],
                                        datosjuridicos["Ente Em."],
                                        datosjuridicos["Fecha registro"],
                                        datosjuridicos["N. Fuente"]
                                        )
                        elif holo:
                            crear_fuente(datosjuridicos["Fuente adm."],
                                        datosjuridicos["Fecha documento"],
                                        datosjuridicos["Ente Em."],
                                        datosjuridicos["Fecha registro"],
                                        datosjuridicos["N. Fuente"]
                                        )

                        time.sleep(1)
                        wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, "body > div.gx-mask.gx-unmask")))
                        driver.find_element(By.CSS_SELECTOR, "#Tab_TAB1Containerpanel4").click()
                        combo = driver.find_element(By.CSS_SELECTOR, "#W0038vACTFUENTADMINID")
                        select = Select(combo)
                        select.select_by_index(1) 
                        archivopropio = indice_folio+str(datos["Folio"])
                        try:
                            time.sleep(1)
                            patron_id = re.compile(r'span_W0038ACTDOCUMENTODESC_\d+')
                            time.sleep(1)
                            elemento1 = driver.find_elements(By.XPATH,"//*[contains(@id, 'span_W0038ACTDOCUMENTODESC_')]")
                            pdfs = len(elemento1)
                            # Obtener el texto de los elementos
                            if archivopropio in elemento1[0].text and pdfs == 2:
                                print ("PDF subidos, ",datos["Folio"])
                            else:
                                borrar_archivos(archivopropio,pdfs)
                                time.sleep(1)
                                subir_documentos(carpeta_almacenamiento,archivopropio)   
                        except: 
                            time.sleep(1)
                            subir_documentos(carpeta_almacenamiento,archivopropio) 
                        
                        lista_cedulas = []
                        lista_bool = []
                        lista_tipoid = []
                        sumatoria = 0
                        firstced = True #no se usa
                        unaveznomas = True               
                        primero = False
                        sininteres = True
                        cedula_sinborrar = ""
                        tipo_idsinborrar = ""  
                            
                        time.sleep(.4)
                        driver.find_element(By.CSS_SELECTOR,'#Tab_TAB1Containerpanel7').click()
                        time.sleep(.4)
                        try:
                            driver.find_element(By.CSS_SELECTOR, "#W0062GRIDPAGINATIONBARContainer_DVPaginationBar > div.PaginationBarCaption.dropdown > div > button").click()
                            driver.find_element(By.XPATH, '//*[@id="W0062GRIDPAGINATIONBARContainer_DVPaginationBar"]/div[2]/div/ul/li[6]/a/span').click()
                            wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, '#W0062GridContainerRow_0001')))   
                        except:
                            print ("Sin interesados registrados, ",datos["Folio"])
                            sininteres = False
                            primero = True  

                        for i in range(cedulas_revisar):  
                            #--------------------------------------------
                            buscar_malos,cedula_sinborrar = buscar_inter_malo(driver, datosjuridicos, sininteres, unaveznomas)
                            #--------------------------------------------
                            if buscar_malos[2]:
                                unaveznomas = False
                                firstced = False
                                if pregunto_derechos == False:
                                    driver.find_element(By.CSS_SELECTOR, "#Tab_TAB1Containerpanel8").click()
                                    try:
                                        espere = WebDriverWait(driver, 2)
                                        espere.until(EC.visibility_of_element_located((By.CSS_SELECTOR, 'img[id^="W0070vDELETE_"]')))
                                        borrar_dere = driver.find_elements(By.CSS_SELECTOR,'img[id^="W0070vDELETE_"]')       #BORRAR DERECHOS
                                        # Itera a través de los botones y haz clic en cada uno de ellos
                                        for boton in borrar_dere:
                                            wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, "body > div.gx-mask.gx-unmask")))
                                            boton.click()
                                    except:
                                        print ("Sin derechos que eliminar, ", datos["Folio"])
                                    pregunto_derechos = True
                                    
                                if sininteres:
                                    panel = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, '#Tab_TAB1Containerpanel6')))
                                    panel.click()
                                    time.sleep(1.5)
                                    wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, 'img[id^="W0054vDELETE2_"]')))
                                    borrar = driver.find_elements(By.CSS_SELECTOR,'img[id^="W0054vDELETE2_"]')         #BORRAR DOCUMENTOS
                                    # Itera a través de los botones y haz clic en cada uno de ellos
                                    for boton in borrar:
                                        wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, "body > div.gx-mask.gx-unmask")))
                                        wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, 'img[id^="W0054vDELETE2_"]')))
                                        time.sleep(1.3)
                                        lolo = driver.find_element(By.CSS_SELECTOR,'img[id^="W0054vDELETE2_"]')
                                        time.sleep(1)
                                        lolo.click()
                                    sininteres = False
                            
                            if any(buscar_malos):
                                r_cedula,r_bool = crear_interes(datosjuridicos["Primer Nombre"],
                                                    datosjuridicos["Segundo Nombre"],
                                                    datosjuridicos["Primer Apellido"],
                                                    datosjuridicos["Segundo Apellido"],
                                                    datosjuridicos["Género"],
                                                    i+1,
                                                    datosjuridicos["Cédulas"]
                                                    )
                                
                                lista_cedulas.append(r_cedula)
                                lista_bool.append(r_bool)
                                lista_tipoid.append(datosjuridicos["Género"])
                            else:
                                lista_cedulas.append(cedula_sinborrar)
                                lista_bool.append(False)
                                lista_tipoid.append(datosjuridicos["Género"])
                            
                            # Extrae los datos de la fila y almacénalos en el diccionario
                            fila_juridico+= 1
                            filaj = hojajuridico[fila_juridico]#tonto
                            for encabezado, columna in datos_titulosjuri.items():
                                if columna:
                                    if filaj[columna-1].value == None:
                                        datosjuridicos[encabezado] = ""
                                    else:
                                        datosjuridicos[encabezado] = filaj[columna-1].value
                            
                        firstced = True
                        i = 0
                        if unaveznomas == False:
                            for ced, ipid in zip(lista_cedulas,lista_tipoid):
                                time.sleep(.5)
                                papu = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, '#Tab_TAB1Containerpanel6')))
                                papu.click()
                                time.sleep(.5)
                                #wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "#span_W0014vACTPREDIOORIGENID")))
                                
                                
                                wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, '#W0054CREARGRUPO')))
                                while True:             #para asegurarse de que hay un grupo creado
                                    try:
                                        driver.find_element(By.CSS_SELECTOR, "#W0054GridgruposContainerRow_0001 > td:nth-child(4) > p")
                                        time.sleep(.5)
                                        grupo = driver.find_element(By.CSS_SELECTOR, "#W0054vSELECCIONAR_0001")      # escoger grupo
                                        time.sleep(.5)
                                        grupo.click()   
                                        time.sleep(.5)
                                        break
                                    except:
                                        driver.find_element(By.CSS_SELECTOR, "#W0054CREARGRUPO").click()
                                        time.sleep(1.5)
                                
                                if ipid != "N":
                                    driver.find_element(By.CSS_SELECTOR, "#W0054vTIPOIDENTIFICACION_ID").send_keys("Cedula Ciudadanía")
                                else:
                                    driver.find_element(By.CSS_SELECTOR, "#W0054vTIPOIDENTIFICACION_ID").send_keys("Numero de Identificación Tributaria")
                                    
                                if firstced:
                                    checkpro = driver.find_element(By.CSS_SELECTOR, "#W0054vACTGRUPINTEREPROPRI")      # escoger grupo
                                    checkpro.click()
                                    firstced = False
                                
                                driver.find_element(By.CSS_SELECTOR, "#W0054vACTMIENINTENUMDOC").send_keys(ced)
                                
                                valor_porcentaje =100/cedulas_revisar
                                valor_porcentajestr = str(valor_porcentaje)
                                valor_porcentajestr = valor_porcentajestr.replace(".",",")
                                
                                participacion = driver.find_element(By.CSS_SELECTOR, "#W0054vACTGRUINTEREPARTIC")
                                participacion.clear()
                                participacion.send_keys(valor_porcentaje)
                                
                                driver.find_elements(By.CSS_SELECTOR, "ody > div.ui-pnotify.stack-topright > div")
                                
                                wait.until(lambda driver: len(driver.find_elements(By.CSS_SELECTOR, ".ui-pnotify.stack-topright[style*='display: block']")) <= 2)
                                driver.find_element(By.CSS_SELECTOR, "#W0054ASOCIARINTERESADO").click()
                                wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, "body > div.gx-mask.gx-unmask"))) 
                                wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "body > div.ui-pnotify.stack-topright > div > div.ui-pnotify-text"))) 
                                time.sleep(1) 
                                count = 0
                                if i == cedulas_revisar - 1:
                                    
                                    while True:
                                        try:
                                            
                                            notificacion = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "body > div.ui-pnotify.stack-topright > div > div.ui-pnotify-text")))  
                                            notitext = notificacion.text
                                            texto = "No es posible agregar el registro ya que la participación total entre los interesados sumaría más de 100..."
                                            texto2 = "Los datos han sido agregados..."
                                            if notificacion == notitext:
                                                wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, "body > div.ui-pnotify.stack-topright > div > div.ui-pnotify-text")))
                                                raise Exception()
                                            else:
                                                break
                                        except:
                                            count +=1
                                            if count >= 6:
                                                print ("ALERTAR, ULTIMO INTERESADO NO PUDO SER INCLUIDO, ",datos["Folio"])
                                                wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, "body > div.ui-pnotify.stack-topright > div > div.ui-pnotify-text")))
                                                break
                                            valor_porcentaje = valor_porcentaje - .1
                                            driver.find_element(By.CSS_SELECTOR, "#W0054vACTGRUINTEREPARTIC").send_keys(str(valor_porcentaje).replace(".",","))
                                            driver.find_element(By.CSS_SELECTOR, "#W0054ASOCIARINTERESADO").click()
                                            wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, "body > div.gx-mask.gx-unmask"))) 
                                sumatoria = sumatoria + valor_porcentaje
                            
                        
                        
                        if not (95 < sumatoria < 100):
                            print ("ALERTA: La suma de los interesados no está entre 95 y 100 % ---> ", sumatoria)
                            
                        
                        if any(lista_bool) and unaveznomas == False:
                            set_zero_intereados(lista_tipoid,lista_cedulas,lista_bool)
                        
                        derechos()
                        
                        time.sleep(.4)
                        driver.find_element(By.CSS_SELECTOR,'#Tab_TAB1Containerpanel7').click()
                        time.sleep(.4)
                        try:
                            driver.find_element(By.CSS_SELECTOR, "#W0062GRIDPAGINATIONBARContainer_DVPaginationBar > div.PaginationBarCaption.dropdown > div > button").click()
                            driver.find_element(By.XPATH, '//*[@id="W0062GRIDPAGINATIONBARContainer_DVPaginationBar"]/div[2]/div/ul/li[6]/a/span').click()
                            wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, '#W0062GridContainerRow_0001')))   
                        except:
                            print ("Sin interesados registrados, ",datos["Folio"])
                            sininteres = False
                            primero = True  

                        
                        for i in range(cedulas_revisar):  
                            #--------------------------------------------
                            if i != 0:
                                fila_juridico_inicial+= 1
                            filaj = hojajuridico[fila_juridico_inicial]#tonto
                            for encabezado, columna in datos_titulosjuri.items():
                                if columna:
                                    if filaj[columna-1].value == None:
                                        datosjuridicos[encabezado] = ""
                                    else:
                                        datosjuridicos[encabezado] = filaj[columna-1].value
                                
                            buscar_malos,poop = buscar_inter_malo(driver, datosjuridicos, True, True)
                            
                            if any(buscar_malos):
                                raise Exception()

                    wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, "body > div.gx-mask.gx-unmask")))
                    driver.find_element(By.CSS_SELECTOR, "#W0070CANCEL").click()
                flag = 0
            archivo_excel.save(carpeta_almacenamiento+"Libro1.xlsx")
            break
    
        except Exception as e:
            print(f"Se produjo un error: {e}")
            a += 1
            flag = 0

            if a == 4:
                error = hoja.cell(row=fila_a_extraer, column=columna_max+ 1)
                error.value = "NO SE LLENÓ FOLIO, REVISAR"
                errorzaso = hoja.cell(row=fila_a_extraer, column=columna_max+ 4)
                errorzaso.value = f"Se produjo un error: {e}"
            
    
    #print ("FOLIO COMPLETADO!!!!!! ----> ", datos["Folio"])
    fila_a_extraer+=1
        


time.sleep(5)
