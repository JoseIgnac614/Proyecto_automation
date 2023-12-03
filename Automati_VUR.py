from selenium import webdriver
from selenium.webdriver.common.by import By  # Importa el módulo By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
import json
import time
import os
import openpyxl
import sys
import io
import pygame
import tkinter as tk
import traceback
from selenium.common.exceptions import NoAlertPresentException

nombrearchivo = "Libro1.xlsx"
#directorio = "C:/Users/nacho/Downloads/davud/Autofinal/CORRECCIOES_PREDIOS_ANTES/Libro1.xlsx"
directorio2 = "C:/Users/nacho/Downloads/Pruebas_autom/QC 01-12-2023/Faltaron/"
directorio = directorio2+nombrearchivo

DirDescargasVUR = 'C:\\Users\\PORTATIL LENOVO\\Downloads\\'
#DirDescargasVUR = 'C:\\Users\\nacho\\Downloads\\'






def reproducir_audio(ruta_audio):
    pygame.init()
    pygame.mixer.init()
    pygame.mixer.music.load(ruta_audio)
    pygame.mixer.music.play()

ruta_audio = "Alarmas/Audio1.mp3"  # Reemplaza con la ruta de tu archivo .mp3

# Configura las opciones de impresión para guardar como PDF
chrome_options = webdriver.ChromeOptions()
chrome_options.add_argument('--kiosk-printing')
chrome_options.add_argument('--disable-gpu')

nivel_de_zoom = 0.8
# Configura las opciones del navegador para ajustar el nivel de zoom
chrome_options.add_argument(f"--force-device-scale-factor={nivel_de_zoom}")


# Configura la opción para guardar como PDF
prefs = {
    'printing.print_preview_sticky_settings.appState': json.dumps({
        'recentDestinations': [{
            'id': 'Save as PDF',
            'origin': 'local',
            'account': '',
        }],
        'selectedDestinationId': 'Save as PDF',
        'version': 2,
    }),
}


# Abre el archivo de Excel
workbook = openpyxl.load_workbook(directorio)

# Selecciona la hoja en la que se encuentra la celda
sheet = workbook['Hoja1']



departameto= 'CORDOBA'
municipio = 'SAHAGUN'           #
circulo = 'ORIP - SAHAGUN'             #ORIP - 
valor_indice = '148'
#valor_excel va a ser el A1 del excel


#print (prefs['download.default_directory'])
# Agrega la preferencia para la ubicación de descarga.
#prefs['download.default_directory'] = 'C:/Users/PORTATIL LENOVO/Downloads/Pruebas_autom'


# Obtener los números del Excel en un vector
numeros = [str(cell.value) for cell in sheet['A'] if cell.value]

# Verificar y eliminar números si existen ambos archivos en el directorio
for numero in numeros.copy():
    nombre_archivo_b = f"148-{numero} B.pdf"
    nombre_archivo_j = f"148-{numero} J.pdf"
    
    ruta_archivo_b = os.path.join(directorio2, nombre_archivo_b)
    ruta_archivo_j = os.path.join(directorio2, nombre_archivo_j)
    
    if os.path.exists(ruta_archivo_b) and os.path.exists(ruta_archivo_j):
        #print(f"Ambos archivos existen para el número {numero}. Se eliminará.")
        numeros.remove(numero)


def mostrar_ventana():
    ventana_emergente = tk.Toplevel(root_window)
    ventana_emergente.title("Mensaje")
    
    label = tk.Label(ventana_emergente, text="¿Ya completó el captcha?")
    label.pack()
    
    button = tk.Button(ventana_emergente, text="Aceptar", command=ventana_emergente.destroy)
    button.pack()
    
    # Obtiene las dimensiones de la pantalla
    screen_width = root_window.winfo_screenwidth()
    screen_height = root_window.winfo_screenheight()
    
    # Calcula la posición para centrar la ventana emergente
    x = (screen_width - ventana_emergente.winfo_reqwidth()) // 2
    y = (screen_height - ventana_emergente.winfo_reqheight()) // 2
    
    # Establece la posición de la ventana emergente
    ventana_emergente.geometry("+{}+{}".format(x, y))
    
    ventana_emergente.transient(root_window)  # Establece la ventana emergente como dependiente de la ventana principal
    ventana_emergente.wait_window(ventana_emergente)  # Espera a que se cierre la ventana emergente


root_window = tk.Tk()
root_window.title("Ventana Principal")


    
def descargar_pdf(driver, download_folder, file_name, nuevo_nombre, tiempo_maximo_espera=30):
    # Realiza la descarga
    time.sleep(2)
    driver.execute_script("window.print();")

    # Inicia un temporizador
    inicio = time.time()

    while time.time() - inicio < tiempo_maximo_espera:
        if os.path.isfile(os.path.join(download_folder, file_name)):
            # El archivo ha sido encontrado, cambia el nombre si es necesario
            if os.path.isfile(os.path.join(download_folder, nuevo_nombre)):
                os.remove(os.path.join(download_folder, nuevo_nombre))  # Elimina el archivo con el nuevo nombre
            if os.path.isfile(os.path.join(directorio2, nuevo_nombre)):
                os.remove(os.path.join(directorio2, nuevo_nombre))  # Elimina el archivo con el nuevo nombre
            os.rename(os.path.join(download_folder, file_name), os.path.join(directorio2, nuevo_nombre))
            
            print("Archivo encontrado y renombrado.")
            return True
        else:
            # El archivo aún no está disponible, espera un segundo y verifica nuevamente
            time.sleep(1)

def wait_n_refresh(tiempo_max_espera, elemento):
    try:
        interaccion = WebDriverWait(driver, tiempo_max_espera).until(EC.presence_of_element_located((By.CSS_SELECTOR, elemento)))
        return interaccion
    except:
        #print("Tiempo de espera agotado. Recargando la página...")
        driver.refresh()

chrome_options.add_experimental_option('prefs', prefs)
timeout = 15

# Inicializa el navegador
driver = webdriver.Chrome(options=chrome_options)


# Maximiza la ventana del navegador
driver.maximize_window()
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
# Navega a la página web deseada
driver.get("https://www.vur.gov.co/siteminderagent/forms_es-ES/loginsnr.fcc?TYPE=100663297&REALMOID=06-6c1363d6-46ce-4693-bea4-501339aa6485&GUID=&SMAUTHREASON=0&METHOD=GET&SMAGENTNAME=-SM-fqc7JfbsitKYA98880nx7GzOWf3PHSx%2frBwpwn0hvw7giR0TRx5Ii32r0m4mIlPP&TARGET=-SM-HTTP%3a%2f%2fwww%2evur%2egov%2eco%2fportal%2fpages%2fvur%2finicio%2ejsf%3furl%3d-%2Fportal-%2FPantallasVUR-%2F-%23-%2F-%3Ftipo-%3DdatosBasicosTierras")

informacion_combobox = {
    '#username': 'LILIANAP.LOPEZ',
    '#password': 'P890980L',
}



for combobox_id, opcion in informacion_combobox.items():
    element = driver.find_element(By.CSS_SELECTOR, combobox_id)
    element.send_keys(opcion)

mostrar_ventana()
        
entrar = driver.find_element(By.CSS_SELECTOR, "#container-login > div > form > div > div.input-submit > input")
entrar.click()


count = 1

while sheet.cell(row=count, column=1).value is not None:
    #hola = sheet.cell(row=count, column=1).value
    if str(sheet.cell(row=count, column=1).value) in numeros:
        contar_malos = 0
        while True:
            try:
                # Lee el valor de la celda
                valor_excel = sheet['A'+str(count)].value  # Reemplaza 'A1' por la celda que deseas usar
                valor_cadena = valor_indice + "-" + str(valor_excel)
                
                # # Esperar 5 segundos
                time.sleep(1)
                driver.switch_to.default_content()   
                entrar = wait_n_refresh(15,"#menu-navegacion > li.dropdown > a")
                entrar.click()

                time.sleep(2)
                # Localiza el elemento sobre el cual deseas pasar el mouse
                elemento = driver.find_element(By.CSS_SELECTOR, "#menu-navegacion > li.dropdown.open > ul > li > a")

                # Crea una instancia de ActionChains
                actions = ActionChains(driver)

                # Mueve el mouse sobre el elemento
                actions.move_to_element(elemento).perform()

                time.sleep(0.5)
                entrar = driver.find_element(By.CSS_SELECTOR, "#menu-navegacion > li.dropdown.open > ul > li > ul > li:nth-child(1) > a")
                entrar.click()

                #time.sleep(5)

                iframe = driver.find_element(By.XPATH, "//iframe[@id='page']")
                driver.switch_to.frame(iframe)

                informacion_combobox = {
                   '#selectDepartamento': departameto,
                   '#selectDepartamento': 'TODOS',
                  '#selectMunicipio': municipio,
                    '#circulo': circulo,
                    '#matricula': valor_excel,
                    # '#criterio':'Referencia Catastral',
                    # '#referenciaCatastral': valor_excel,
                }                  

                entrar = wait_n_refresh(10,"#selectDepartamento")

                flag = 0                    #Para avisar que el folio no está en el VU
                start_time = time.time()
                break
            except:
                driver.refresh()
        wait2 = WebDriverWait(driver, 50)
        while True:
            try:
                if flag != 2:
                    # Llena la información en los combobox en la página actual
                    for combobox_id, opcion in informacion_combobox.items():
                        element = driver.find_element(By.CSS_SELECTOR, combobox_id)
                        element.send_keys(opcion)
                        time.sleep(.8)

                    wait = WebDriverWait(driver, timeout)
                    # Haz clic en el botón "Continuar" (ajusta el selector según tu página)
                    matricula = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, 'body > div.wrapper.ng-scope > div > div:nth-child(2) > div.panel-body > div.btn-group.pull-right > button')))
                    matricula.click()
                    time.sleep(1.5)    
                    #Espera hasta que aparezca el elemento en la siguiente página
                    

                    # Esperar hasta que el elemento tenga display: none
                                        
                    
                    try:
                        #wait2.until(driver.execute_script("return window.getComputedStyle(document.querySelector('#esperaModal')).getPropertyValue('display')") != 'block')
                        wait2.until(EC.invisibility_of_element_located((By.ID, "esperaModal")))
                        #print("El elemento con id='esperaModal' tiene display: none")

                    except:
                        try:
                            alert = driver.switch_to.alert
                            #print("Alerta encontrada:", alert.text)
                            alert.accept()  # Para aceptar la alerta
                            matricula.click()
                            
                        except NoAlertPresentException:
                            print("")
                    
                    
                    matricula = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "body > div.wrapper.ng-scope > div > div:nth-child(3) > div.panel-body > table > tbody > tr > td:nth-child(1) > a")))
                    matricula.click()
                    #Si el elemento se encuentra, el bucle se detiene
                    
                    try:
                        #wait2.until(driver.execute_script("return window.getComputedStyle(document.querySelector('#esperaModal')).getPropertyValue('display')") != 'block')
                        wait2.until(EC.invisibility_of_element_located((By.ID, "esperaModal")))
                        #print("El elemento con id='esperaModal' tiene display: none")

                    except:
                        try:
                            alert = driver.switch_to.alert
                            #print("Alerta encontrada:", alert.text)
                            alert.accept()  # Para aceptar la alerta
                        except NoAlertPresentException:
                            print("")
                    
                    
                    matricula = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, 'div.panel.panel-primary.datos-basicos[ng-show="pantallaDatosBasicos"]:not(.ng-hide)'))
                    )

                    # Abre el diálogo de impresión del navegador
                    descargar_pdf(driver,DirDescargasVUR,"-VUR.pdf",valor_cadena+" B.pdf")
                    # Encuentra todos los elementos que tienen un atributo "id"
                    driver.switch_to.default_content()

                        
                    entrar = driver.find_element(By.CSS_SELECTOR, "#menu-navegacion > li.dropdown > a")
                    entrar.click()
                    time.sleep(0.5)
                    # Localiza el elemento sobre el cual deseas pasar el mouse
                    elemento = driver.find_element(By.CSS_SELECTOR, "#menu-navegacion > li.dropdown.open > ul > li > a")

                    # Crea una instancia de ActionChains
                    actions = ActionChains(driver)

                    # Mueve el mouse sobre el elemento
                    actions.move_to_element(elemento).perform()

                    time.sleep(0.5)
                    entrar = driver.find_element(By.CSS_SELECTOR, "#menu-navegacion > li.dropdown.open > ul > li > ul > li:nth-child(2) > a")
                    entrar.click()

                    
                    
                    iframe = driver.find_element(By.XPATH, "//iframe[@id='page']")
                    driver.switch_to.frame(iframe)
                    flag = 2
                time.sleep(2)
                WebDriverWait(driver, timeout).until(
                    EC.visibility_of_element_located((By.CSS_SELECTOR, 'div.panel.panel-primary[ng-show="pantallaDatosJuridicos"]:not(.ng-hide)'))
                )

                try:
                    #wait2.until(driver.execute_script("return window.getComputedStyle(document.querySelector('#esperaModal')).getPropertyValue('display')") != 'block')
                    wait2.until(EC.invisibility_of_element_located((By.ID, "esperaModal")))
                    #print("El elemento con id='esperaModal' tiene display: none")

                except:
                    try:
                        alert = driver.switch_to.alert
                        #print("Alerta encontrada:", alert.text)
                        alert.accept()  # Para aceptar la alerta
                    except NoAlertPresentException:
                        print("")
            
                # Esperar hasta que se encuentre un elemento <li> que contiene "Lista"
                elemento = WebDriverWait(driver, timeout).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, "body > div.wrapper.ng-scope > div > div:nth-child(11) > div > div.tabbable.ng-isolate-scope > ul > li:nth-child(2)"))
                )
                elemento.click()
                descargar_pdf(driver,DirDescargasVUR,"-VUR.pdf",valor_cadena+" J.pdf")
                
                
                edito = driver.find_element(By.CSS_SELECTOR, "body > div.wrapper.ng-scope > div > div.panel.panel-primary.panel-botones > div > div > div > button:nth-child(3)")
                
                driver.execute_script("arguments[0].scrollIntoView();", edito)   
                time.sleep(.5)     
        
                encontrado = False
                conta = 0
                vayase = False
                while not encontrado:
                    try:
                        # Intenta encontrar el elemento
                        edito.click()
                        encontrado = True
                    except:
                        
                        # Si no se encuentra el elemento, desplaza el scroll hacia arriba
                        driver.execute_script("window.scrollBy(0, -100);")
                        # Espera un momento para que la página se cargue y se actualice
                        time.sleep(.7)
                        if conta > 8:
                            vayase = True
                            break
                        conta += 1
                if vayase == True:
                    raise Exception("ERROR: Se dañó en el ciclo del while de interesados") 
                
                driver.switch_to.default_content()
                
                break

            except Exception as e:
                #print(f"Se produjo un error: {e}")
                # Si el elemento no se encuentra, recarga la página
                mensaje_error = traceback.format_exc()
                try:
                    driver.switch_to.alert.accept()
                except:
                    print("")
                #esperaModal
                driver.refresh()
                time.sleep(3)
                iframe = driver.find_element(By.XPATH, "//iframe[@id='page']")
                driver.switch_to.frame(iframe)
                
                
                elapsed_time = time.time() - start_time                 #Por si el folio no está en el vur
                if elapsed_time >= 10:
                    if flag == 2:
                        start_time = time.time()
                        flag = 1
                    else:
                        flag = 1
                        contar_malos += 1
                        if contar_malos > 8:
                            time.sleep(1)
                            edito = driver.find_element(By.CSS_SELECTOR, "body > div.wrapper.ng-scope > div > div.panel.panel-primary.panel-botones > div > div > div > button:nth-child(3)")                
                            time.sleep(1)
                            edito2 = driver.find_element(By.CSS_SELECTOR, "body > div.wrapper.ng-scope > div > div:nth-child(2) > div.panel-body > div.btn-group.pull-right > button")                
                            
                            driver.execute_script("arguments[0].scrollIntoView();", edito)   
                            time.sleep(.5)     
                    
                            encontrado = False
                            conta = 0
                            vayase = False
                            while not encontrado:
                                try:
                                    time.sleep(1)
                                    # Intenta encontrar el elemento
                                    edito.click()
                                    encontrado = True
                                except:
                                    
                                    # Si no se encuentra el elemento, desplaza el scroll hacia arriba
                                    driver.execute_script("window.scrollBy(0, -100);")
                                    # Espera un momento para que la página se cargue y se actualice
                                    time.sleep(.7)
                                    if conta > 8:
                                        if vayase == False:
                                            vayase = True
                                            conta = 0
                                            edito = edito2  
                                        else:
                                            break
                                                                                                                      
                                    conta += 1
                            if encontrado == False:
                                raise Exception("ERROR: Se dañó en el ciclo del while de interesados") 
                            break
                    
                
        if flag == 1:
            #escribir algo en el excel
            #print ("No está el folio en el VUR")
            sheet['B'+str(count)] = "No está el folio en el VUR"
            workbook.save(directorio)
            flag = 0
            try:
                driver.switch_to.alert.accept()  # Para aceptar la alerta (hacer clic en "Aceptar")
            except:
                print ("")
            driver.switch_to.default_content()
            wait = WebDriverWait(driver, timeout)
            #matricula = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "#menu-navegacion > li:nth-child(5) > a"))).click()
    
    # Obtener la lista de archivos en el directorio
    archivos = os.listdir(DirDescargasVUR)

    # Eliminar archivos que contienen 'VUR' en su nombre
    for archivo in archivos:
        if 'VUR' in archivo:
            os.remove(os.path.join(DirDescargasVUR, archivo))      
    count += 1
#root_window.mainloop()
reproducir_audio(ruta_audio)
time.sleep(30)

